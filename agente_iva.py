# -*- coding: utf-8 -*-
"""
agente_iva.py — Motor principal del Agente IVA Devoluciones SAT
Procesa CFDIs, estado de cuenta bancario y auxiliar SAP para generar
4 entregables de auditoría fiscal.
"""

import os
import sys
import json

# Forzar UTF-8 en stdout para evitar errores de encoding en Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import re
import copy
import datetime
import xml.etree.ElementTree as ET
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdfplumber
import fitz  # pymupdf
from docx import Document
from docx.shared import Pt, Cm, RGBColor
from docx.enum.text import WD_ALIGN_PARAGRAPH
from num2words import num2words

# ─── Variables de configuración ────────────────────────────────────────────
TOLERANCIA_MONTO = 1.00       # diferencia máxima en pesos para cruce
TOLERANCIA_DIAS  = 3          # días de diferencia permitidos
PREFIJO_CRUCE    = "CRZ"      # prefijo del ID de cruce

# ─── Namespaces SAT ────────────────────────────────────────────────────────
NS = {
    "cfdi":   "http://www.sat.gob.mx/cfd/4",
    "pago20": "http://www.sat.gob.mx/Pagos20",
    "tfd":    "http://www.sat.gob.mx/TimbreFiscalDigital",
}

# ─── Mapa formas de pago ───────────────────────────────────────────────────
FORMAS_PAGO = {
    "01": "Efectivo",
    "02": "Cheque nominativo",
    "03": "Transferencia electrónica",
    "04": "Tarjeta de crédito",
    "05": "Monedero electrónico",
    "06": "Dinero electrónico",
    "08": "Vales de despensa",
    "12": "Dación en pago",
    "13": "Pago por subrogación",
    "14": "Pago por consignación",
    "15": "Condonación",
    "17": "Compensación",
    "23": "Novación",
    "24": "Confusión",
    "25": "Remisión de deuda",
    "26": "Prescripción o caducidad",
    "27": "A satisfacción del acreedor",
    "28": "Tarjeta de débito",
    "29": "Tarjeta de servicios",
    "30": "Aplicación de anticipos",
    "31": "Intermediario pagos",
    "99": "Por definir",
}

# ─── Colores Excel ─────────────────────────────────────────────────────────
COLOR_OK_BG    = "E2EFDA"
COLOR_OK_FG    = "375623"
COLOR_WARN_BG  = "FFF2CC"
COLOR_WARN_FG  = "7F6000"
COLOR_ERR_BG   = "FCE4D6"
COLOR_ERR_FG   = "C00000"
COLOR_IVA_BG   = "EBF5E0"


def progreso(paso: str, pct: int, msg: str):
    """Emite línea de progreso para que ui.py la capture."""
    print(f"PROGRESO:{paso}:{pct}:{msg}", flush=True)


def resultado(total_cfdis, iva_confirmado, cruces_completos, sin_cruce,
              iva_trasladado=0.0, iva_acreditable=0.0, saldo_favor=0.0):
    """Emite línea de resultado final."""
    print(f"RESULTADO:{total_cfdis}:{iva_confirmado:.2f}:{cruces_completos}:{sin_cruce}:{iva_trasladado:.2f}:{iva_acreditable:.2f}:{saldo_favor:.2f}", flush=True)


def error(msg: str):
    """Emite línea de error."""
    print(f"ERROR:{msg}", flush=True)


# ══════════════════════════════════════════════════════════════════════════════
# PASO 1 — CARGA DE CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════

def cargar_config(base_dir: Path) -> dict:
    """Carga config.json. Crea vacío si no existe."""
    cfg_path = base_dir / "input" / "config.json"
    campos = ["empresa", "rfc", "domicilio", "clabe", "rep_legal", "rfc_rep",
              "autorizados", "folio_sat"]
    if not cfg_path.exists():
        cfg = {c: "" for c in campos}
        with open(cfg_path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        print("ADVERTENCIA: config.json no encontrado, se creó vacío. "
              "Complétalo antes de generar el escrito.", flush=True)
        return cfg
    with open(cfg_path, encoding="utf-8") as f:
        cfg = json.load(f)
    # Agregar campos faltantes
    for c in campos:
        if c not in cfg:
            cfg[c] = ""
    return cfg


# ══════════════════════════════════════════════════════════════════════════════
# PASO 2 — PARSEO DE CFDIs COMPLEMENTO DE PAGO
# ══════════════════════════════════════════════════════════════════════════════

def parsear_cfdis(base_dir: Path, carpeta: str = "cfdi",
                  tipo_cfdi: str = "pago") -> tuple[list, list]:
    """
    Lee todos los XMLs en input/{carpeta}/.
    Retorna (registros, errores).
    Cada registro es un dict con datos del pago y sus DoctoRelacionados.
    """
    cfdi_dir = base_dir / "input" / carpeta
    if not cfdi_dir.exists():
        return [], []
    archivos = list(cfdi_dir.glob("*.xml")) + list(cfdi_dir.glob("*.XML"))
    registros = []
    errores = []
    total = len(archivos)

    for i, xml_path in enumerate(archivos):
        pct = int((i + 1) / max(total, 1) * 100)
        progreso("cfdi", pct, f"Procesando XML {i+1} de {total}")
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()

            # Verificar tipo de comprobante
            tipo = root.get("TipoDeComprobante", "")
            if tipo != "P":
                errores.append({
                    "archivo": xml_path.name,
                    "error": f"TipoDeComprobante={tipo} (se esperaba P)",
                    "fecha": datetime.datetime.now().isoformat(),
                })
                continue

            # ── Datos del comprobante raíz ───────────────────────────────
            fecha_emision = root.get("Fecha", "")
            emisor_node   = root.find("cfdi:Emisor", NS)
            receptor_node = root.find("cfdi:Receptor", NS)
            tfd_node      = root.find(".//tfd:TimbreFiscalDigital", NS)

            uuid_cp = tfd_node.get("UUID", "") if tfd_node is not None else ""
            rfc_emisor      = emisor_node.get("Rfc", "")   if emisor_node is not None else ""
            nombre_emisor   = emisor_node.get("Nombre", "") if emisor_node is not None else ""
            rfc_receptor    = receptor_node.get("Rfc", "") if receptor_node is not None else ""
            nombre_receptor = receptor_node.get("Nombre", "") if receptor_node is not None else ""

            # ── Nodo Pago ────────────────────────────────────────────────
            for pago_node in root.findall(".//pago20:Pago", NS):
                fecha_pago    = pago_node.get("FechaPago", "")[:10]
                forma_pago_c  = pago_node.get("FormaDePagoP", "")
                forma_pago    = FORMAS_PAGO.get(forma_pago_c, forma_pago_c)
                monto_pago    = float(pago_node.get("Monto", "0") or "0")
                moneda_p      = pago_node.get("MonedaP", "MXN")
                num_operacion = pago_node.get("NumOperacion", "").strip()
                banco_ord     = pago_node.get("NomBancoOrdExt", "")
                cta_ord       = pago_node.get("CtaOrdenante", "")
                cta_ben       = pago_node.get("CtaBeneficiario", "")

                # ── DoctoRelacionado ─────────────────────────────────────
                for docto in pago_node.findall("pago20:DoctoRelacionado", NS):
                    uuid_factura  = docto.get("IdDocumento", "")
                    serie         = docto.get("Serie", "")
                    folio         = docto.get("Folio", "")
                    serie_folio   = f"{serie}-{folio}".strip("-")
                    moneda_doc    = docto.get("MonedaDR", "MXN")
                    tipo_cambio   = float(docto.get("TipoCambioDR", "1") or "1")
                    metodo_pago   = docto.get("MetodoDePagoDR", "")
                    parcialidad   = docto.get("NumParcialidad", "")
                    saldo_ant     = float(docto.get("ImpSaldoAnt", "0") or "0")
                    imp_pagado    = float(docto.get("ImpPagado", "0") or "0")
                    saldo_ins     = float(docto.get("ImpSaldoInsoluto", "0") or "0")

                    # IVA 16% — los atributos en Pagos 2.0 tienen sufijo "DR"
                    iva16 = 0.0
                    for tras in docto.findall(".//pago20:TrasladoDR", NS):
                        # Soporte para ambas variantes de nombre de atributo
                        tasa = (tras.get("TasaOCuotaDR") or
                                tras.get("TasaOCuota") or "")
                        impo = (tras.get("ImporteDR") or
                                tras.get("Importe") or "0")
                        if tasa in ("0.160000", "0.16"):
                            iva16 += float(impo or "0")

                    # IVA Retenido
                    iva_ret = 0.0
                    for ret in docto.findall(".//pago20:RetencionDR", NS):
                        impo = (ret.get("ImporteDR") or
                                ret.get("Importe") or "0")
                        iva_ret += float(impo or "0")

                    # Convertir a MXN si es necesario
                    if moneda_doc != "MXN" and tipo_cambio:
                        imp_pagado_mxn = imp_pagado * tipo_cambio
                        iva16_mxn      = iva16 * tipo_cambio
                    else:
                        imp_pagado_mxn = imp_pagado
                        iva16_mxn      = iva16

                    # Cuenta propia: CtaBeneficiario en cobros, CtaOrdenante en pagos
                    if tipo_cfdi == "cobro":
                        cuenta_nuestra = cta_ben  # recibimos en cta_ben
                    else:
                        cuenta_nuestra = cta_ord  # pagamos desde cta_ord

                    registros.append({
                        "uuid_cp":          uuid_cp,
                        "fecha_emision":    fecha_emision[:10],
                        "fecha_pago":       fecha_pago,
                        "forma_pago":       forma_pago,
                        "num_operacion":    num_operacion,
                        "banco_ord":        banco_ord,
                        "cta_ord":          cta_ord,
                        "cta_ben":          cta_ben,
                        "cuenta_nuestra":   cuenta_nuestra,
                        "tipo_cfdi":        tipo_cfdi,
                        "rfc_emisor":       rfc_emisor,
                        "nombre_emisor":    nombre_emisor,
                        "rfc_receptor":     rfc_receptor,
                        "nombre_receptor":  nombre_receptor,
                        "uuid_factura":     uuid_factura,
                        "serie_folio":      serie_folio,
                        "moneda_doc":       moneda_doc,
                        "tipo_cambio":      tipo_cambio,
                        "metodo_pago":      metodo_pago,
                        "parcialidad":      parcialidad,
                        "saldo_anterior":   saldo_ant,
                        "importe_pagado":   imp_pagado,
                        "importe_pagado_mxn": imp_pagado_mxn,
                        "saldo_insoluto":   saldo_ins,
                        "iva16":            iva16,
                        "iva16_mxn":        iva16_mxn,
                        "iva_retenido":     iva_ret,
                        "monto_pago":       monto_pago,
                        "moneda_p":         moneda_p,
                        # Campos de cruce (se llenarán después)
                        "id_cruce":         "",
                        "cruce_edo_cuenta": False,
                        "cruce_sap":        False,
                        "metodo_cruce":     "",
                        "observaciones":    "",
                    })

        except Exception as e:
            errores.append({
                "archivo": xml_path.name,
                "error": str(e),
                "fecha": datetime.datetime.now().isoformat(),
            })

    progreso("cfdi", 100, f"{len(registros)} registros extraídos de {total} XMLs")
    return registros, errores


# ══════════════════════════════════════════════════════════════════════════════
# PASO 3 — LECTURA DEL ESTADO DE CUENTA (PDF)
# ══════════════════════════════════════════════════════════════════════════════

def _parsear_monto(texto: str) -> float:
    """Convierte texto de monto (1,234.56) a float."""
    if not texto:
        return 0.0
    limpio = re.sub(r"[^0-9.\-]", "", texto.replace(",", ""))
    try:
        return float(limpio)
    except ValueError:
        return 0.0


def _extraer_referencia(texto: str) -> str:
    """Extrae referencia numérica de 7–18 dígitos de un texto."""
    matches = re.findall(r"\b\d{7,18}\b", texto)
    return matches[0] if matches else ""


def _nuevo_movimiento(fecha, desc, ref, cargo, abono, saldo):
    return {"fecha": fecha, "descripcion": desc, "referencia": ref,
            "cargo": cargo, "abono": abono, "saldo": saldo,
            "banco": "", "cuenta_bancaria": "", "moneda": "MXN", "archivo_pdf": "",
            "id_cruce": "", "cruce_cfdi": False,
            "uuid_cfdi": "", "metodo_cruce": ""}


# Meses abreviados en español para PDFs bancarios MX
_MESES_ABR = {
    "ENE": "01", "FEB": "02", "MAR": "03", "ABR": "04",
    "MAY": "05", "JUN": "06", "JUL": "07", "AGO": "08",
    "SEP": "09", "OCT": "10", "NOV": "11", "DIC": "12",
}


def _normalizar_fecha_banco(texto: str, año_default: int) -> str:
    """Convierte '09/ENE' o '09/01' o '09/01/2026' a 'YYYY-MM-DD'."""
    texto = texto.strip()
    # Formato DD/MMM (BBVA: 09/ENE)
    m = re.match(r"^(\d{1,2})[/\-]([A-Za-z]{3})$", texto)
    if m:
        dia = m.group(1).zfill(2)
        mes = _MESES_ABR.get(m.group(2).upper(), "01")
        return f"{año_default}-{mes}-{dia}"
    # Formato DD/MM/YYYY
    m = re.match(r"^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})$", texto)
    if m:
        dia, mes, año = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        if len(año) == 2:
            año = "20" + año
        return f"{año}-{mes}-{dia}"
    return texto


def _parsear_bloque_bbva(lineas: list, año: int) -> list:
    """
    Parser específico para estado de cuenta BBVA MX (texto extraído de pdfplumber).
    Formato observado:
      DD/MMM DD/MMM COD  DESCRIPCION  CARGO_O_ABONO
      REFERENCIA Ref. DESCRIPCION_2
      ...
      (líneas de detalle adicionales)
    """
    movs = []
    # Regex para línea principal con fecha
    PAT_MOV = re.compile(
        r"^(\d{1,2}/[A-Za-z]{3})\s+\d{1,2}/[A-Za-z]{3}\s+\w+\s+"
        r"(.+?)\s+([\d,]+\.\d{2})\s*$"
    )
    # Regex para línea de referencia numérica (10-20 dígitos)
    PAT_REF  = re.compile(r"^(\d{7,20})\s*(Ref\.|$)")

    i = 0
    while i < len(lineas):
        linea = lineas[i].strip()
        m = PAT_MOV.match(linea)
        if m:
            fecha_raw = m.group(1)
            desc      = m.group(2).strip()
            monto     = _parsear_monto(m.group(3))
            fecha     = _normalizar_fecha_banco(fecha_raw, año)
            # Buscar referencia en la siguiente línea
            ref = ""
            if i + 1 < len(lineas):
                sig = lineas[i + 1].strip()
                mr  = PAT_REF.match(sig)
                if mr:
                    ref = mr.group(1)
                    # Agregar descripcion de la línea de referencia
                    desc = desc + " " + sig
                    i += 1
                else:
                    ref = _extraer_referencia(sig)
                    if ref:
                        desc = desc + " " + sig
                        i += 1
            movs.append(_nuevo_movimiento(fecha, desc[:120], ref, 0.0, monto, 0.0))
        i += 1
    return movs


_MESES_ES_NUM = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}

_ENCABEZADOS_TABLA = {
    "FECHA", "FOLIO", "DESCRIPCION", "DEPOSITO", "RETIRO", "SALDO",
    "PRODUCTO", "CUENTA", "ANTERIOR", "ANTERIOR:", "FINAL:", "PERIODO",
    "NUMERO", "DE", "MONEDA", "RFC", "SUCURSAL", "INTERESES", "ISR",
    "COMISIONES", "GAT", "NOMINAL", "REAL", "RETENIDO", "NETOS", "BRUTOS",
}


def _extraer_cuenta_bancaria(texto: str, banco: str) -> str:
    """
    Extrae el número de cuenta bancaria del texto de la primera página del PDF.
    Soporta formatos Santander (65-XXXXXXXX-X), BBVA (10-11 dígitos), y genérico.
    """
    t = texto.upper()
    # Santander: XX-XXXXXXXX-X
    m = re.search(r'\b(\d{2}-\d{8}-\d{1})\b', texto)
    if m:
        return m.group(1)
    # BBVA / genérico: buscar "CUENTA" seguido de número
    m = re.search(r'(?:CUENTA|NO\.?|NUMERO|N[UÚ]MERO)[^\d]{0,20}(\d{10,18})', t)
    if m:
        return m.group(1)
    # CLABE (18 dígitos)
    m = re.search(r'\b(\d{18})\b', texto)
    if m:
        return m.group(1)
    # Número de cuenta (10-11 dígitos)
    m = re.search(r'\b(\d{10,11})\b', texto)
    if m:
        return m.group(1)
    return ""


def _detectar_banco(texto: str) -> str:
    """Detecta el banco a partir del texto de la primera página del PDF."""
    t = texto.upper()
    if "BBVA" in t or "BANCOMER" in t:
        return "BBVA"
    if "SANTANDER" in t:
        return "SANTANDER"
    if "CITIBANAMEX" in t or "BANAMEX" in t:
        return "BANAMEX"
    if "BANORTE" in t:
        return "BANORTE"
    if "HSBC" in t:
        return "HSBC"
    if "SCOTIABANK" in t:
        return "SCOTIABANK"
    if "INBURSA" in t:
        return "INBURSA"
    return "DESCONOCIDO"


def _parsear_santander_words(words: list, año: int) -> list:
    """
    Parser Santander MX usando coordenadas de palabras (fitz get_text("words")).
    Formato: DD-MMM-YYYY | FOLIO | DESCRIPCION (multi-línea) | DEPOSITO | RETIRO | SALDO
    Usa posición X para distinguir DEPOSITO vs RETIRO.
    """
    if not words:
        return []

    # Determinar ancho útil de la página
    max_x = max(w[2] for w in words) if words else 600.0

    # Umbrales de columnas Santander (% del ancho)
    x_deposito = max_x * 0.60
    x_retiro   = max_x * 0.75
    x_saldo    = max_x * 0.88

    # Agrupar palabras por línea (y0 con tolerancia ±3pt)
    from collections import defaultdict
    lineas_dict: dict = defaultdict(list)
    for w in words:
        x0, y0, x1, y1, text = w[0], w[1], w[2], w[3], w[4]
        y_key = round(y0 / 3) * 3
        lineas_dict[y_key].append((x0, text))

    pat_fecha  = re.compile(
        r'^(\d{1,2})-(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)-(\d{4})$',
        re.IGNORECASE,
    )
    pat_monto  = re.compile(r'^[\d,]+\.\d{2}$')
    pat_folio  = re.compile(r'^\d{1,7}$')

    movs    = []
    current = None

    def _guardar(c):
        if c:
            movs.append(_nuevo_movimiento(
                c["fecha"], c["desc"][:200],
                c["ref"], c["cargo"], c["abono"], c["saldo"],
            ))

    for y in sorted(lineas_dict.keys()):
        linea = sorted(lineas_dict[y], key=lambda w: w[0])
        if not linea:
            continue

        first_x, first_word = linea[0]
        m_fecha = pat_fecha.match(first_word)

        if m_fecha:
            _guardar(current)
            dia  = int(m_fecha.group(1))
            mes  = _MESES_ES_NUM.get(m_fecha.group(2).upper(), 1)
            anio = int(m_fecha.group(3))
            try:
                fecha = datetime.date(anio, mes, dia).isoformat()
            except Exception:
                fecha = f"{anio}-{mes:02d}-{dia:02d}"

            current = {"fecha": fecha, "ref": "", "desc": "",
                       "cargo": 0.0, "abono": 0.0, "saldo": 0.0}

            for x, word in linea[1:]:
                if pat_monto.match(word):
                    monto = _parsear_monto(word)
                    if x >= x_saldo:
                        current["saldo"] = monto
                    elif x >= x_retiro:
                        current["cargo"] = monto
                    elif x >= x_deposito:
                        current["abono"] = monto
                    elif not current["ref"] and pat_folio.match(word):
                        current["ref"] = word
                    else:
                        current["desc"] = (current["desc"] + " " + word).strip()
                elif not current["ref"] and pat_folio.match(word):
                    current["ref"] = word
                else:
                    current["desc"] = (current["desc"] + " " + word).strip()

        elif current is not None:
            # Línea de continuación — agregar a descripción / completar montos
            for x, word in linea:
                if pat_monto.match(word):
                    monto = _parsear_monto(word)
                    if x >= x_saldo and current["saldo"] == 0.0:
                        current["saldo"] = monto
                    elif x >= x_retiro and current["cargo"] == 0.0:
                        current["cargo"] = monto
                    elif x >= x_deposito and current["abono"] == 0.0:
                        current["abono"] = monto
                else:
                    if word.upper() not in _ENCABEZADOS_TABLA:
                        current["desc"] = (current["desc"] + " " + word).strip()[:200]

    _guardar(current)
    return movs


def _parsear_texto_generico(texto_pagina: str, año: int) -> list:
    """
    Parser de línea por línea para bancos con formato desconocido.
    Extrae cualquier línea con fecha + monto.
    """
    movs = []
    lineas = texto_pagina.split("\n")
    for linea in lineas:
        linea = linea.strip()
        if not linea:
            continue
        # Formato DD/MM/YYYY
        fm = re.search(r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})", linea)
        # Formato DD/MMM (para BBVA sin match del bloque)
        if not fm:
            fm = re.search(r"(\d{1,2}/[A-Za-z]{3})", linea)
        if not fm:
            continue
        fecha_raw = fm.group(1)
        fecha = _normalizar_fecha_banco(fecha_raw, año)
        montos_enc = re.findall(r"[\d,]+\.\d{2}", linea)
        cargo_val = abono_val = 0.0
        if len(montos_enc) >= 2:
            cargo_val = _parsear_monto(montos_enc[-2])
            abono_val = _parsear_monto(montos_enc[-1])
        elif len(montos_enc) == 1:
            abono_val = _parsear_monto(montos_enc[-1])
        ref = _extraer_referencia(linea)
        movs.append(_nuevo_movimiento(fecha, linea[:120], ref,
                                      cargo_val, abono_val, 0.0))
    return movs


def leer_estado_cuenta(base_dir: Path) -> list:
    """
    Lee PDFs en input/pdf_bancos/ (o input/estado_cuenta/ legacy).
    Retorna lista de movimientos con campos:
      fecha, descripcion, referencia, cargo, abono, saldo, cruce_*
    """
    pdf_dir = base_dir / "input" / "pdf_bancos"
    if not pdf_dir.exists() or not (list(pdf_dir.glob("*.pdf")) + list(pdf_dir.glob("*.PDF"))):
        pdf_dir = base_dir / "input" / "estado_cuenta"
    pdfs    = list(pdf_dir.glob("*.pdf")) + list(pdf_dir.glob("*.PDF"))
    movimientos = []
    año_default = datetime.datetime.now().year

    for pdf_path in pdfs:
        progreso("estado_cuenta", 10, f"Leyendo {pdf_path.name}...")
        try:
            doc     = fitz.open(str(pdf_path))
            n_pages = len(doc)

            # Detectar banco y año en primera página
            primera = doc[0].get_text("text", sort=True) if n_pages > 0 else ""
            banco   = _detectar_banco(primera)
            m_año   = re.search(r"(20\d{2})", primera)
            if m_año:
                año_default = int(m_año.group(1))

            es_bbva      = banco == "BBVA" or bool(
                re.search(r"\d{1,2}/[A-Z]{3}\s+\d{1,2}/[A-Z]{3}", primera))
            es_santander = banco == "SANTANDER"

            movs_pdf = []
            for i, page in enumerate(doc):
                pct = 10 + int((i + 1) / max(n_pages, 1) * 80)
                progreso("estado_cuenta", pct,
                         f"{banco} {pdf_path.name} pag {i+1}/{n_pages}")

                if es_santander:
                    words = page.get_text("words", sort=True)
                    movs_pagina = _parsear_santander_words(words, año_default)
                elif es_bbva:
                    texto  = page.get_text("text", sort=True) or ""
                    lineas = [l for l in texto.split("\n") if l.strip()]
                    movs_pagina = _parsear_bloque_bbva(lineas, año_default)
                else:
                    texto = page.get_text("text", sort=True) or ""
                    if not texto.strip():
                        continue
                    movs_pagina = _parsear_texto_generico(texto, año_default)

                movs_pdf.extend(movs_pagina)

            # Extraer número de cuenta bancaria del encabezado
            cuenta_pdf = _extraer_cuenta_bancaria(primera, banco)

            # Detectar moneda del estado de cuenta
            t_up = primera.upper()
            if any(k in t_up for k in ("DOLAR", "USD", "DOLARES", "DIVISA")):
                moneda_pdf = "USD"
            else:
                moneda_pdf = "MXN"

            # Etiquetar con banco, cuenta y moneda de origen
            for mv in movs_pdf:
                mv["banco"]           = banco
                mv["cuenta_bancaria"] = cuenta_pdf
                mv["moneda"]          = moneda_pdf
                mv["archivo_pdf"]     = pdf_path.name

            movimientos.extend(movs_pdf)
            doc.close()
            print(f"PROGRESO:estado_cuenta:90:{banco} {pdf_path.name}: "
                  f"{len(movs_pdf)} movimientos", flush=True)

        except Exception as e:
            print(f"ADVERTENCIA: No se pudo leer {pdf_path.name}: {e}", flush=True)

    progreso("estado_cuenta", 100, f"{len(movimientos)} movimientos extraídos del PDF")
    return movimientos


# ══════════════════════════════════════════════════════════════════════════════
# PASO 4 — LECTURA DEL AUXILIAR SAP
# ══════════════════════════════════════════════════════════════════════════════

# Mapeo de nombres de columna SAP (varios idiomas)
_COL_MAP = {
    "fecha":         ["fecha contable", "fecha doc", "budat", "buchungsdatum",
                      "posting date", "fecha documento",
                      "fecha contabiliz.", "fecha contabilizacion",
                      "fecha de documento", "fecha contabilización"],
    "num_documento": ["nº documento", "numero documento", "belegnummer", "belnr",
                      "document number", "doc.num",
                      "n° documento", "no documento", "ndocumento",
                      "número de documento"],
    "referencia":    ["referencia", "nº referencia", "xblnr", "reference",
                      "referencia doc.", "referenz",
                      "asignación", "asignacion", "assignment",
                      "referencia banco", "ref banco"],
    "concepto":      ["texto", "descripción", "descripcion", "text", "sgtxt",
                      "posting text"],
    "importe":       ["importe en ml", "importe", "betrag", "amount", "monto",
                      "importe ml",
                      "importe en moneda local", "importe moneda local",
                      "amount in local currency"],
    "sentido":       ["haber", "debe", "s/h", "debit/credit", "s_h", "d/c",
                      "clave contabiliz.", "clave contabilizacion",
                      "posting key"],
    "cuenta":        ["cuenta", "cuenta mayor", "konto", "account", "gl account",
                      "cuenta gl"],
    "sociedad":      ["sociedad", "bukrs", "company code", "sociedad co."],
    "importe_me":    ["importe en ml2", "importe ml2", "importe me",
                      "importe moneda extranjera", "importe en moneda extranjera",
                      "amount in foreign currency", "hwbas"],
    "moneda_fuerte": ["mon.moneda fuerte", "moneda fuerte", "foreign currency",
                      "moneda extranjera", "waers", "currency key",
                      "mon. moneda fuerte"],
}


def _detectar_columna(df_cols: list, candidatos: list) -> str | None:
    """Retorna el nombre real de la columna que coincide con algún candidato."""
    cols_lower = {c.lower().strip(): c for c in df_cols}
    for cand in candidatos:
        if cand.lower() in cols_lower:
            return cols_lower[cand.lower()]
    return None


def leer_auxiliar_sap(base_dir: Path,
                      carpeta: str = "aux_pagado") -> tuple[pd.DataFrame | None, dict, list]:
    """
    Lee el Excel de auxiliar SAP.
    carpeta: 'aux_pagado' (IVA acreditable) o 'aux_cobrado' (IVA trasladado)
    Retorna (df_original, col_map, advertencias).
    """
    def _xls(d): return (list(d.glob("*.xlsx")) + list(d.glob("*.xls")) +
                         list(d.glob("*.XLSX")) + list(d.glob("*.XLS")))
    aux_dir  = base_dir / "input" / carpeta
    archivos = _xls(aux_dir)
    if not archivos:
        # Legacy fallback
        aux_dir  = base_dir / "input" / "auxiliar"
        archivos = _xls(aux_dir)
    if not archivos:
        return None, {}, [f"No se encontro archivo Excel en {carpeta}/ ni auxiliar/"]

    advertencias = []
    dfs = []

    for arch in archivos:
        try:
            progreso("auxiliar_sap", 10, f"Leyendo {arch.name}...")
            # Intentar leer desde la primera hoja, tolerando encabezados sucios
            df_raw = pd.read_excel(arch, header=None, dtype=str)
            # Buscar fila de encabezado
            header_row = 0
            for idx, row in df_raw.iterrows():
                vals = [str(v).lower().strip() for v in row if pd.notna(v)]
                keywords = ["fecha", "importe", "cuenta", "belnr", "budat", "belegnummer"]
                if any(k in " ".join(vals) for k in keywords):
                    header_row = idx
                    break
            df_arch = pd.read_excel(arch, header=header_row, dtype=str)
            df_arch.columns = [str(c).strip() for c in df_arch.columns]
            dfs.append(df_arch)
        except Exception as e:
            advertencias.append(f"Error leyendo {arch.name}: {e}")

    if not dfs:
        return None, {}, advertencias

    # Concatenar todos los Excel (pueden tener columnas ligeramente distintas)
    try:
        df = pd.concat(dfs, ignore_index=True)
    except Exception:
        df = dfs[0]

    # Mapear columnas
    col_map = {}
    for campo, candidatos in _COL_MAP.items():
        col = _detectar_columna(list(df.columns), candidatos)
        if col:
            col_map[campo] = col
        else:
            advertencias.append(f"Columna '{campo}' no detectada en SAP "
                                 f"(buscando: {candidatos[:3]}...)")

    return df, col_map, advertencias


def leer_auxiliar_bancos(base_dir: Path) -> dict:
    """
    Lee el auxiliar SAP de bancos.
    Retorna dict: {cuenta_bancaria_normalizada: {cuenta_sap, nombre_banco, ...}}
    Soporta columnas: CUENTA SAP, NOMBRE BANCO, CUENTA BANCARIA (o CUENTA BA)
    """
    def _xls(d):
        return (list(d.glob("*.xlsx")) + list(d.glob("*.xls")) +
                list(d.glob("*.XLSX")) + list(d.glob("*.XLS")))

    aux_dir  = base_dir / "input" / "aux_bancos"
    archivos = _xls(aux_dir)
    if not archivos:
        return {}

    cuentas = {}
    for arch in archivos:
        try:
            df_raw = pd.read_excel(arch, header=None, dtype=str)
            # Buscar fila de encabezado
            header_row = 0
            for idx, row in df_raw.iterrows():
                vals = " ".join(str(v).lower() for v in row if pd.notna(v))
                if "cuenta" in vals and ("banco" in vals or "sap" in vals):
                    header_row = idx
                    break
            df = pd.read_excel(arch, header=header_row, dtype=str)
            df.columns = [str(c).strip().upper() for c in df.columns]

            # Detectar columnas flexiblemente
            col_sap    = next((c for c in df.columns if "SAP" in c or c == "CUENTA"), None)
            col_banco  = next((c for c in df.columns if "BANCO" in c and "CUENTA" not in c), None)
            col_cuenta = next((c for c in df.columns if "BANCARIA" in c or "CUENTA BA" in c
                               or (c.startswith("CUENTA") and c != col_sap)), None)

            if not col_cuenta:
                continue

            for _, row in df.iterrows():
                cta_raw = str(row.get(col_cuenta, "") or "").strip()
                if not cta_raw or cta_raw.lower() in ("nan", "none", ""):
                    continue
                # Normalizar: quitar guiones y espacios para comparación
                cta_norm = re.sub(r'[\s\-]', '', cta_raw)
                cuentas[cta_norm] = {
                    "cuenta_bancaria": cta_raw,
                    "cuenta_sap":  str(row.get(col_sap, "") or "").strip(),
                    "nombre_banco": str(row.get(col_banco, "") or "").strip(),
                }
        except Exception as e:
            print(f"ADVERTENCIA auxiliar bancos {arch.name}: {e}", flush=True)

    return cuentas


def _normalizar_cuenta(cuenta: str) -> str:
    """Quita guiones, espacios y ceros iniciales para comparación flexible."""
    return re.sub(r'[\s\-]', '', cuenta).lstrip("0")


# ══════════════════════════════════════════════════════════════════════════════
# LÓGICA DE CRUCE
# ══════════════════════════════════════════════════════════════════════════════

def _fecha_a_date(texto: str) -> datetime.date | None:
    """Convierte varios formatos de fecha a date."""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y",
                "%Y%m%d", "%d.%m.%Y"):
        try:
            return datetime.datetime.strptime(texto.strip(), fmt).date()
        except Exception:
            pass
    return None


def cruzar_con_banco(registros: list, movimientos: list, periodo_str: str) -> int:
    """
    Cruza registros de CFDI con movimientos bancarios.
    - Agrupa movimientos por cuenta bancaria para evitar falsos cruces
    - Respeta moneda: USD vs MXN
    - Modifica en lugar registros y movimientos.
    Retorna contador de cruces.
    """
    cruce_seq   = [0]
    total_cruce = 0

    def nuevo_id():
        cruce_seq[0] += 1
        return f"{PREFIJO_CRUCE}-{periodo_str}-{cruce_seq[0]:04d}"

    # Agrupar movimientos por cuenta (normalizada) y por moneda
    idx_ref_global: dict = {}
    movs_por_cuenta: dict = {}
    for m in movimientos:
        cta = _normalizar_cuenta(m.get("cuenta_bancaria", "") or "")
        movs_por_cuenta.setdefault(cta, []).append(m)
        if m["referencia"]:
            idx_ref_global.setdefault(m["referencia"], []).append(m)

    def _candidatos_para_registro(r) -> list:
        """
        Devuelve movimientos candidatos de la misma cuenta bancaria.
        - COBRO: cuenta_nuestra = CtaBeneficiario (recibimos)
        - PAGO:  cuenta_nuestra = CtaOrdenante   (pagamos)
        Si está vacío o no hay match, devuelve todos los movimientos.
        """
        cta_cfdi = _normalizar_cuenta(r.get("cuenta_nuestra", "") or "")
        if cta_cfdi:
            # Buscar coincidencia parcial (últimos 4+ dígitos)
            for cta_key in movs_por_cuenta:
                if cta_key.endswith(cta_cfdi) or cta_cfdi.endswith(cta_key):
                    return movs_por_cuenta[cta_key]
        # Fallback: todos los movimientos
        return movimientos

    def _monto_movimiento(m: dict, moneda_r: str) -> float:
        """Usa abono/cargo según moneda; para USD usa cargo si la cuenta es extranjera."""
        if moneda_r == "MXN":
            return m["abono"] if m["abono"] else m["cargo"]
        else:
            # Para moneda extranjera, el monto puede estar en cargo o abono
            return m["abono"] if m["abono"] else m["cargo"]

    for r in registros:
        if r["cruce_edo_cuenta"]:
            continue

        num_op     = r["num_operacion"]
        fecha_p    = _fecha_a_date(r["fecha_pago"])
        moneda_r   = r.get("moneda_doc", "MXN")
        # Para USD usar importe_pagado (moneda original); para MXN usar MXN
        monto_r    = (r["importe_pagado"] if moneda_r != "MXN"
                      else r["importe_pagado_mxn"])
        candidatos = _candidatos_para_registro(r)
        id_asignar = ""

        # Nivel 1: EXACTO por num_operacion
        if num_op:
            for m in idx_ref_global.get(num_op, []):
                if not m["cruce_cfdi"]:
                    id_asignar = nuevo_id()
                    m.update(id_cruce=id_asignar, cruce_cfdi=True,
                             uuid_cfdi=r["uuid_cp"], metodo_cruce="EXACTO")
                    r["metodo_cruce"] = "EXACTO"
                    break

        # Nivel 2: Monto + Fecha (dentro de candidatos de la misma cuenta)
        if not id_asignar:
            for m in candidatos:
                if m["cruce_cfdi"]:
                    continue
                monto_m = _monto_movimiento(m, moneda_r)
                tol = max(1.0, monto_r * 0.005)
                if abs(monto_m - monto_r) <= tol:
                    fecha_m = _fecha_a_date(m["fecha"])
                    if fecha_p and fecha_m and abs((fecha_p - fecha_m).days) <= TOLERANCIA_DIAS:
                        id_asignar = nuevo_id()
                        m.update(id_cruce=id_asignar, cruce_cfdi=True,
                                 uuid_cfdi=r["uuid_cp"], metodo_cruce="MONTO+FECHA")
                        r["metodo_cruce"] = "MONTO+FECHA"
                        break

        # Nivel 3: Solo monto (cruce débil, busca en todos)
        if not id_asignar:
            for m in movimientos:
                if m["cruce_cfdi"]:
                    continue
                monto_m = _monto_movimiento(m, moneda_r)
                tol = max(1.0, monto_r * 0.005)
                if abs(monto_m - monto_r) <= tol:
                    id_asignar = nuevo_id()
                    m.update(id_cruce=id_asignar, cruce_cfdi=True,
                             uuid_cfdi=r["uuid_cp"], metodo_cruce="SOLO_MONTO")
                    r["metodo_cruce"] = "SOLO_MONTO (débil)"
                    break

        if id_asignar:
            r["id_cruce"]         = id_asignar
            r["cruce_edo_cuenta"] = True
            total_cruce += 1

    return total_cruce


def cruzar_con_sap(registros: list, df_sap: pd.DataFrame | None,
                   col_map: dict, periodo_str: str) -> int:
    """
    Cruza registros de CFDI con filas del auxiliar SAP.
    Agrega columnas id_cruce, uuid_cfdi, estado_cruce al df_sap.
    """
    if df_sap is None:
        return 0

    df_sap["id_cruce_sap"]   = ""
    df_sap["uuid_cfdi_sap"]  = ""
    df_sap["estado_cruce"]   = "Sin cruce"

    col_ref       = col_map.get("referencia", "")
    col_bel       = col_map.get("num_documento", "")
    col_imp       = col_map.get("importe", "")       # Importe en moneda local (MXN)
    col_imp_me    = col_map.get("importe_me", "")    # Importe en moneda extranjera (USD)
    col_mon_fuerte= col_map.get("moneda_fuerte", "") # Columna Mon.Moneda fuerte
    col_fec       = col_map.get("fecha", "")

    cruce_seq = [1000]  # secuencia SAP separada para evitar colisiones

    def nuevo_id_sap():
        cruce_seq[0] += 1
        return f"{PREFIJO_CRUCE}-{periodo_str}-{cruce_seq[0]:04d}"

    total_cruce = 0

    # Columna adicional de asignación (puede ser referencia de banco)
    col_asig = None
    for cname in df_sap.columns:
        if "asignaci" in cname.lower() or "asignacion" in cname.lower():
            col_asig = cname
            break

    for r in registros:
        num_op    = r["num_operacion"]
        moneda_r  = r.get("moneda_doc", "MXN")
        monto_r   = (r["importe_pagado"] if moneda_r != "MXN"
                     else r["importe_pagado_mxn"])
        iva_r     = r["iva16_mxn"]
        fecha_p   = _fecha_a_date(r["fecha_pago"])

        tol_din = max(TOLERANCIA_MONTO, monto_r * 0.005)
        tol_iva = max(TOLERANCIA_MONTO, iva_r   * 0.005)

        for idx, fila in df_sap.iterrows():
            if df_sap.at[idx, "estado_cruce"] == "Cruzado":
                continue

            ref_sap  = str(fila.get(col_ref,  "")).strip() if col_ref   else ""
            bel_sap  = str(fila.get(col_bel,  "")).strip() if col_bel   else ""
            asig_sap = str(fila.get(col_asig, "")).strip() if col_asig  else ""
            fec_str  = str(fila.get(col_fec,  "")).strip()  if col_fec  else ""

            # Seleccionar columna de importe según moneda de la fila SAP
            mon_fila = str(fila.get(col_mon_fuerte, "MXN") or "MXN").strip().upper()
            if col_imp_me and mon_fila not in ("", "MXN", "NAN", "NONE"):
                # Cuenta en moneda extranjera: usar Importe ML2
                imp_str = str(fila.get(col_imp_me, "0") or "0").strip()
            else:
                imp_str = str(fila.get(col_imp, "0") or "0").strip() if col_imp else "0"

            imp_sap  = _parsear_monto(imp_str.replace(",", "."))
            fecha_s  = _fecha_a_date(fec_str)

            # Normalizar asignación: quitar sufijo de año (ej. "15000001172026 *" → "1500000117")
            asig_norm = re.sub(r"20\d{2}.*$", "", asig_sap.replace(" ", "")).strip()

            # Nivel 1: Exacto por referencia, BELNR o asignación == NumOperacion
            if num_op and (ref_sap == num_op or bel_sap == num_op
                           or asig_sap.startswith(num_op) or asig_norm == num_op):
                id_uso = r["id_cruce"] if r["id_cruce"] else nuevo_id_sap()
                df_sap.at[idx, "id_cruce_sap"]  = id_uso
                df_sap.at[idx, "uuid_cfdi_sap"] = r["uuid_cp"]
                df_sap.at[idx, "estado_cruce"]  = "Cruzado"
                r["cruce_sap"] = True
                if not r["id_cruce"]:
                    r["id_cruce"] = id_uso
                total_cruce += 1
                break

            # Nivel 2a: IVA del CFDI == importe SAP + Fecha
            if iva_r > 0 and abs(abs(imp_sap) - iva_r) <= tol_iva:
                if fecha_p and fecha_s:
                    if abs((fecha_p - fecha_s).days) <= TOLERANCIA_DIAS:
                        id_uso = r["id_cruce"] if r["id_cruce"] else nuevo_id_sap()
                        df_sap.at[idx, "id_cruce_sap"]  = id_uso
                        df_sap.at[idx, "uuid_cfdi_sap"] = r["uuid_cp"]
                        df_sap.at[idx, "estado_cruce"]  = "Cruzado"
                        r["cruce_sap"] = True
                        if not r["id_cruce"]:
                            r["id_cruce"] = id_uso
                        total_cruce += 1
                        break

            # Nivel 2b: Monto total del CFDI == importe SAP + Fecha
            if abs(abs(imp_sap) - monto_r) <= tol_din:
                if fecha_p and fecha_s:
                    if abs((fecha_p - fecha_s).days) <= TOLERANCIA_DIAS:
                        id_uso = r["id_cruce"] if r["id_cruce"] else nuevo_id_sap()
                        df_sap.at[idx, "id_cruce_sap"]  = id_uso
                        df_sap.at[idx, "uuid_cfdi_sap"] = r["uuid_cp"]
                        df_sap.at[idx, "estado_cruce"]  = "Cruzado"
                        r["cruce_sap"] = True
                        if not r["id_cruce"]:
                            r["id_cruce"] = id_uso
                        total_cruce += 1
                        break

        # Nivel 3: Solo IVA o monto (cruce débil)
        if not r["cruce_sap"]:
            for idx, fila in df_sap.iterrows():
                if df_sap.at[idx, "estado_cruce"] != "Sin cruce":
                    continue
                imp_str = str(fila.get(col_imp, "0")).strip() if col_imp else "0"
                imp_sap = _parsear_monto(imp_str.replace(",", "."))
                if (iva_r > 0 and abs(abs(imp_sap) - iva_r) <= tol_iva) or \
                   abs(abs(imp_sap) - monto_r) <= tol_din:
                    id_uso = r["id_cruce"] if r["id_cruce"] else nuevo_id_sap()
                    df_sap.at[idx, "id_cruce_sap"]  = id_uso
                    df_sap.at[idx, "uuid_cfdi_sap"] = r["uuid_cp"]
                    df_sap.at[idx, "estado_cruce"]  = "Cruce debil"
                    r["cruce_sap"] = True
                    if not r["id_cruce"]:
                        r["id_cruce"] = id_uso
                    break

    progreso("auxiliar_sap", 100, f"Auxiliar SAP cruzado: {total_cruce} coincidencias")
    return total_cruce


# ══════════════════════════════════════════════════════════════════════════════
# PASO 5 — EXCEL DE REPORTE
# ══════════════════════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Arial")


def _borde_delgado() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _color_fila(r: dict) -> tuple[str, str]:
    """Retorna (bg, fg) según estado de cruce."""
    if r["cruce_edo_cuenta"] and r["cruce_sap"]:
        return COLOR_OK_BG, COLOR_OK_FG
    elif r["cruce_edo_cuenta"] or r["cruce_sap"]:
        return COLOR_WARN_BG, COLOR_WARN_FG
    else:
        return COLOR_ERR_BG, COLOR_ERR_FG


def generar_excel(registros: list, movimientos: list,
                  df_sap: pd.DataFrame | None, errores: list,
                  base_dir: Path, periodo_str: str,
                  tipo: str = "acreditable") -> Path:
    """
    Genera el Excel de reporte.
    tipo='acreditable' → IVA Pagado a Proveedores
    tipo='trasladado'  → IVA Cobrado a Clientes
    """
    label   = "IVA_Acreditable" if tipo == "acreditable" else "IVA_Trasladado"
    out_path = base_dir / "output" / f"reporte_{label}_{periodo_str}.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── HOJA 1: Detalle (nombre según tipo) ──────────────────────────────
    hoja1_nombre = "Detalle Cobros" if tipo == "trasladado" else "Detalle Pagos"
    ws1 = wb.create_sheet(hoja1_nombre)
    encabezados = [
        "ID Cruce", "UUID Complemento", "Fecha Emisión CP", "Fecha Pago",
        "Forma de Pago", "Num. Operación Banco", "Banco Ordenante",
        "Cta. Ordenante", "Cta. Beneficiario", "RFC Emisor", "Nombre Emisor",
        "RFC Receptor", "Nombre Receptor", "UUID Factura Relacionada",
        "Serie-Folio", "Parcialidad", "Subtotal Pagado", "IVA 16%",
        "IVA Retenido", "Total Pagado", "Moneda", "Tipo de Cambio",
        "Saldo Anterior", "Saldo Insoluto",
        "✓ Edo. Cuenta", "✓ Auxiliar SAP", "Método Cruce", "Observaciones",
    ]
    # Encabezado
    for col, enc in enumerate(encabezados, 1):
        cell = ws1.cell(row=1, column=col, value=enc)
        cell.fill  = _fill("1F4E79")
        cell.font  = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _borde_delgado()

    for i, r in enumerate(registros, 2):
        bg, fg = _color_fila(r)
        datos = [
            r["id_cruce"], r["uuid_cp"], r["fecha_emision"], r["fecha_pago"],
            r["forma_pago"], r["num_operacion"], r["banco_ord"],
            r["cta_ord"], r["cta_ben"], r["rfc_emisor"], r["nombre_emisor"],
            r["rfc_receptor"], r["nombre_receptor"], r["uuid_factura"],
            r["serie_folio"], r["parcialidad"],
            r["importe_pagado_mxn"], r["iva16_mxn"], r["iva_retenido"],
            r["importe_pagado_mxn"] + r["iva16_mxn"],
            r["moneda_doc"], r["tipo_cambio"],
            r["saldo_anterior"], r["saldo_insoluto"],
            "✓" if r["cruce_edo_cuenta"] else "✗",
            "✓" if r["cruce_sap"] else "✗",
            r["metodo_cruce"], r["observaciones"],
        ]
        for col, val in enumerate(datos, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            # IVA columns (17, 18, 19) → fondo verde claro
            if col in (17, 18, 19):
                cell.fill = _fill(COLOR_IVA_BG)
            else:
                cell.fill = _fill(bg)
            cell.font = _font(color=fg)
            cell.border = _borde_delgado()
            if col in (17, 18, 19, 20, 22, 23, 24):
                cell.number_format = '"$"#,##0.00'
            elif col in (3, 4):
                cell.number_format = "DD/MM/YYYY"

    # Congelar primera fila y auto-filtros
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions
    # Anchos de columna
    anchos = [14, 36, 12, 12, 22, 20, 22, 18, 18, 14, 28, 14, 28,
              36, 12, 10, 14, 12, 12, 14, 8, 10, 14, 14, 6, 6, 16, 30]
    for col, ancho in enumerate(anchos, 1):
        ws1.column_dimensions[get_column_letter(col)].width = ancho
    ws1.row_dimensions[1].height = 28

    # ── HOJA 2: Resumen Ejecutivo ─────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen Ejecutivo")
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 20

    def _res_enc(txt):
        cell = ws2.cell(row=ws2.max_row + 1, column=1, value=txt)
        cell.fill  = _fill("1F4E79")
        cell.font  = _font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="left")
        ws2.merge_cells(start_row=cell.row, start_column=1,
                        end_row=cell.row, end_column=2)

    def _res_fila(label, val):
        r_idx = ws2.max_row + 1
        ws2.cell(row=r_idx, column=1, value=label).font = _font(size=10)
        c = ws2.cell(row=r_idx, column=2, value=val)
        c.font = _font(bold=True, size=10)
        c.alignment = Alignment(horizontal="right")

    total_cfdis    = len(set(r["uuid_cp"] for r in registros))
    total_facturas = len(registros)
    iva_total      = sum(r["iva16_mxn"] for r in registros)
    iva_confirma   = sum(r["iva16_mxn"] for r in registros if r["cruce_edo_cuenta"])
    cruce_completo = sum(1 for r in registros if r["cruce_edo_cuenta"] and r["cruce_sap"])
    cruce_parcial  = sum(1 for r in registros if r["cruce_edo_cuenta"] != r["cruce_sap"])
    sin_cruce      = sum(1 for r in registros if not r["cruce_edo_cuenta"] and not r["cruce_sap"])

    _res_enc("RESUMEN EJECUTIVO — AGENTE IVA DEVOLUCIONES SAT")
    _res_fila("Periodo procesado", periodo_str)
    _res_fila("Fecha y hora de generación",
              datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    ws2.append([])
    _res_enc("VOLUMEN")
    _res_fila("Total CFDIs Complemento de Pago", total_cfdis)
    _res_fila("Total facturas relacionadas (DoctoRelacionado)", total_facturas)
    ws2.append([])
    _res_enc("IVA")
    _res_fila("IVA 16% total en CFDIs", f"${iva_total:,.2f}")
    _res_fila("SALDO A FAVOR A SOLICITAR (cruce edo. cuenta)", f"${iva_confirma:,.2f}")
    ws2.append([])
    _res_enc("CRUCES")
    pct_cobertura = (cruce_completo / max(total_facturas, 1)) * 100
    _res_fila("Cruces completos (CFDI + banco + SAP)", cruce_completo)
    _res_fila("Cruce parcial (solo banco o solo SAP)", cruce_parcial)
    _res_fila("Sin cruce", sin_cruce)
    _res_fila("% Cobertura de cruce", f"{pct_cobertura:.1f}%")

    # ── HOJA 3: Estado de Cuenta ──────────────────────────────────────────
    ws3 = wb.create_sheet("Estado de Cuenta")
    enc3 = ["ID Cruce", "Fecha", "Descripción", "Referencia",
            "Cargo", "Abono", "Saldo", "✓ CFDI", "UUID Complemento", "Método Cruce"]
    for col, enc in enumerate(enc3, 1):
        c = ws3.cell(row=1, column=col, value=enc)
        c.fill = _fill("1F4E79")
        c.font = _font(bold=True, color="FFFFFF")
        c.border = _borde_delgado()
    for i, m in enumerate(movimientos, 2):
        bg = COLOR_OK_BG if m["cruce_cfdi"] else COLOR_ERR_BG
        fg = COLOR_OK_FG if m["cruce_cfdi"] else COLOR_ERR_FG
        datos = [m["id_cruce"], m["fecha"], m["descripcion"], m["referencia"],
                 m["cargo"], m["abono"], m["saldo"],
                 "✓" if m["cruce_cfdi"] else "✗",
                 m["uuid_cfdi"], m["metodo_cruce"]]
        for col, val in enumerate(datos, 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.fill   = _fill(bg)
            c.font   = _font(color=fg)
            c.border = _borde_delgado()
            if col in (5, 6, 7):
                c.number_format = '"$"#,##0.00'
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions

    # ── HOJA 4: Auxiliar SAP ──────────────────────────────────────────────
    ws4 = wb.create_sheet("Auxiliar SAP")
    if df_sap is not None:
        cols_orig = list(df_sap.columns)
        enc4 = cols_orig + ["ID Cruce", "UUID CFDI", "Estado Cruce"]
        for col, enc in enumerate(enc4, 1):
            c = ws4.cell(row=1, column=col, value=enc)
            c.fill = _fill("1F4E79")
            c.font = _font(bold=True, color="FFFFFF")
            c.border = _borde_delgado()
        for i, (_, row) in enumerate(df_sap.iterrows(), 2):
            estado = str(row.get("estado_cruce", "Sin cruce"))
            if estado == "Cruzado":
                bg, fg = COLOR_OK_BG, COLOR_OK_FG
            elif estado == "Cruce débil":
                bg, fg = COLOR_WARN_BG, COLOR_WARN_FG
            else:
                bg, fg = COLOR_ERR_BG, COLOR_ERR_FG
            for col, campo in enumerate(cols_orig, 1):
                c = ws4.cell(row=i, column=col, value=row.get(campo, ""))
                c.fill   = _fill(bg)
                c.font   = _font(color=fg)
                c.border = _borde_delgado()
            n = len(cols_orig)
            for col, campo in enumerate(["id_cruce_sap", "uuid_cfdi_sap", "estado_cruce"], n + 1):
                c = ws4.cell(row=i, column=col, value=row.get(campo, ""))
                c.fill   = _fill(bg)
                c.font   = _font(color=fg, bold=(campo == "estado_cruce"))
                c.border = _borde_delgado()
        ws4.freeze_panes = "A2"
        ws4.auto_filter.ref = ws4.dimensions

    # ── HOJA 5: Sin Cruce ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("Sin Cruce")
    enc5 = ["Origen", "Fecha", "Referencia", "Monto",
            "Descripción", "Posible causa", "Acción sugerida"]
    for col, enc in enumerate(enc5, 1):
        c = ws5.cell(row=1, column=col, value=enc)
        c.fill = _fill("C00000")
        c.font = _font(bold=True, color="FFFFFF")
    fila_sc = 2
    for r in registros:
        if not r["cruce_edo_cuenta"]:
            causa = "Pago sin CFDI emitido" if not r["num_operacion"] else "CFDI de periodo anterior"
            ws5.cell(row=fila_sc, column=1, value="CFDI")
            ws5.cell(row=fila_sc, column=2, value=r["fecha_pago"])
            ws5.cell(row=fila_sc, column=3, value=r["num_operacion"])
            ws5.cell(row=fila_sc, column=4, value=r["importe_pagado_mxn"]).number_format = '"$"#,##0.00'
            ws5.cell(row=fila_sc, column=5, value=r["nombre_emisor"])
            ws5.cell(row=fila_sc, column=6, value=causa)
            ws5.cell(row=fila_sc, column=7, value="Verificar con proveedor y banco")
            fila_sc += 1
    for m in movimientos:
        if not m["cruce_cfdi"] and (m["abono"] or m["cargo"]):
            ws5.cell(row=fila_sc, column=1, value="Banco")
            ws5.cell(row=fila_sc, column=2, value=m["fecha"])
            ws5.cell(row=fila_sc, column=3, value=m["referencia"])
            monto = m["abono"] or m["cargo"]
            ws5.cell(row=fila_sc, column=4, value=monto).number_format = '"$"#,##0.00'
            ws5.cell(row=fila_sc, column=5, value=m["descripcion"][:60])
            ws5.cell(row=fila_sc, column=6, value="Movimiento bancario sin reflejo contable")
            ws5.cell(row=fila_sc, column=7, value="Solicitar CFDI al proveedor")
            fila_sc += 1

    # ── HOJA 6: Errores y Advertencias ───────────────────────────────────
    ws6 = wb.create_sheet("Errores y Advertencias")
    enc6 = ["Tipo", "Archivo / Detalle", "Error", "Fecha"]
    for col, enc in enumerate(enc6, 1):
        c = ws6.cell(row=1, column=col, value=enc)
        c.fill = _fill("7F7F7F")
        c.font = _font(bold=True, color="FFFFFF")
    for i, err in enumerate(errores, 2):
        ws6.cell(row=i, column=1, value="XML")
        ws6.cell(row=i, column=2, value=err.get("archivo", ""))
        ws6.cell(row=i, column=3, value=err.get("error", ""))
        ws6.cell(row=i, column=4, value=err.get("fecha", ""))

    wb.save(out_path)
    progreso("excel", 100,
             f"Excel generado con {len(registros)} filas en {out_path.name}")
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
# PASO 6 — MARCADO DEL PDF
# ══════════════════════════════════════════════════════════════════════════════

def marcar_pdf(movimientos: list, base_dir: Path, periodo_str: str) -> Path | None:
    """Copia el PDF del estado de cuenta y agrega marcas de cruce en margen derecho."""
    pdf_dir = base_dir / "input" / "pdf_bancos"
    if not pdf_dir.exists() or not (list(pdf_dir.glob("*.pdf")) + list(pdf_dir.glob("*.PDF"))):
        pdf_dir = base_dir / "input" / "estado_cuenta"
    pdfs    = list(pdf_dir.glob("*.pdf")) + list(pdf_dir.glob("*.PDF"))
    if not pdfs:
        progreso("pdf", 100, "No se encontró PDF para marcar")
        return None

    src_pdf  = pdfs[0]
    out_path = base_dir / "output" / f"estado_cuenta_cruzado_{periodo_str}.pdf"

    doc = fitz.open(str(src_pdf))

    # Colores
    verde = (0, 100 / 255, 0)
    rojo  = (180 / 255, 0, 0)

    for page in doc:
        ancho = page.rect.width
        for m in movimientos:
            # Buscar por referencia o monto en el texto de la página
            texto_buscar = m["referencia"] if m["referencia"] else ""
            instancias = []
            if texto_buscar:
                instancias = page.search_for(texto_buscar)

            if instancias:
                rect_base = instancias[0]
                x_margin  = ancho - 55
                y_centro  = (rect_base.y0 + rect_base.y1) / 2

                if m["cruce_cfdi"]:
                    etiqueta = m["id_cruce"]
                    color    = verde
                else:
                    etiqueta = "S/C"
                    color    = rojo

                # Fondo blanco semitransparente
                rect_txt = fitz.Rect(x_margin - 2, y_centro - 5,
                                     ancho - 2, y_centro + 7)
                page.draw_rect(rect_txt, color=(1, 1, 1), fill=(1, 1, 1),
                               fill_opacity=0.85)

                # Texto del ID de cruce
                page.insert_text(
                    (x_margin, y_centro + 4),
                    etiqueta,
                    fontname="helv",
                    fontsize=7,
                    color=color,
                )

                # Línea delgada
                if m["cruce_cfdi"]:
                    page.draw_line(
                        (rect_base.x1 + 2, y_centro),
                        (x_margin - 2, y_centro),
                        color=verde,
                        width=0.5,
                    )

    # Página final de resumen
    pagina_res = doc.new_page(width=612, height=792)
    y = 40
    pagina_res.insert_text((40, y), "RESUMEN DE AUDITORÍA DE CRUCES",
                            fontname="helv", fontsize=14,
                            color=(0.12, 0.31, 0.47))
    y += 25
    pagina_res.insert_text(
        (40, y),
        f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  |  "
        f"Periodo: {periodo_str}",
        fontname="helv", fontsize=9, color=(0.4, 0.4, 0.4),
    )
    y += 20

    # Encabezado de tabla
    enc_col = [40, 110, 200, 340, 490]
    enc_txt = ["ID Cruce", "Fecha", "Referencia Banco", "UUID CFDI", "Monto"]
    for cx, txt in zip(enc_col, enc_txt):
        pagina_res.insert_text((cx, y), txt, fontname="helv",
                               fontsize=8, color=(1, 1, 1))
    pagina_res.draw_rect(
        fitz.Rect(35, y - 11, 580, y + 3),
        color=(0.12, 0.31, 0.47),
        fill=(0.12, 0.31, 0.47),
    )
    for cx, txt in zip(enc_col, enc_txt):
        pagina_res.insert_text((cx, y), txt, fontname="helv",
                               fontsize=8, color=(1, 1, 1))
    y += 14

    cruzados = [m for m in movimientos if m["cruce_cfdi"]]
    for m in cruzados[:60]:  # máximo 60 filas para no saturar
        if y > 760:
            break
        datos = [m["id_cruce"], m["fecha"],
                 m["referencia"][:18] if m["referencia"] else "",
                 m["uuid_cfdi"][:30] if m["uuid_cfdi"] else "",
                 f'${m["abono"] or m["cargo"]:,.2f}']
        for cx, txt in zip(enc_col, datos):
            pagina_res.insert_text((cx, y), str(txt),
                                   fontname="helv", fontsize=7.5,
                                   color=(0, 0, 0))
        y += 12

    doc.save(str(out_path))
    progreso("pdf", 100, f"PDF marcado con {len(cruzados)} referencias")
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
# PASO 7 — AUXILIAR SAP CRUZADO
# ══════════════════════════════════════════════════════════════════════════════

def guardar_auxiliar_sap_cruzado(df_sap: pd.DataFrame | None,
                                  base_dir: Path, periodo_str: str,
                                  tipo: str = "acreditable") -> Path | None:
    """Guarda el auxiliar SAP con columnas de cruce y colores.
    tipo='acreditable' → auxiliar_IVA_Acreditable_*.xlsx
    tipo='trasladado'  → auxiliar_IVA_Trasladado_*.xlsx
    """
    if df_sap is None:
        progreso("auxiliar_cruzado", 100, "No hay auxiliar SAP para guardar")
        return None

    label    = "IVA_Acreditable" if tipo == "acreditable" else "IVA_Trasladado"
    out_path = base_dir / "output" / f"auxiliar_{label}_{periodo_str}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auxiliar Cruzado"

    cols_orig = [c for c in df_sap.columns
                 if c not in ("id_cruce_sap", "uuid_cfdi_sap", "estado_cruce")]
    enc_todos = cols_orig + ["ID Cruce", "UUID CFDI", "Estado Cruce"]

    for col, enc in enumerate(enc_todos, 1):
        c = ws.cell(row=1, column=col, value=enc)
        c.fill = _fill("1F4E79")
        c.font = _font(bold=True, color="FFFFFF")
        c.border = _borde_delgado()

    for i, (_, row) in enumerate(df_sap.iterrows(), 2):
        estado = str(row.get("estado_cruce", "Sin cruce"))
        if estado == "Cruzado":
            bg, fg = COLOR_OK_BG, COLOR_OK_FG
        elif estado == "Cruce débil":
            bg, fg = COLOR_WARN_BG, COLOR_WARN_FG
        else:
            bg, fg = COLOR_ERR_BG, COLOR_ERR_FG

        for col, campo in enumerate(cols_orig, 1):
            c = ws.cell(row=i, column=col, value=row.get(campo, ""))
            c.fill = _fill(bg); c.font = _font(color=fg); c.border = _borde_delgado()
        n = len(cols_orig)
        for col, campo in enumerate(["id_cruce_sap", "uuid_cfdi_sap", "estado_cruce"], n + 1):
            c = ws.cell(row=i, column=col, value=row.get(campo, ""))
            c.fill = _fill(bg); c.font = _font(color=fg, bold=(col == n + 3))
            c.border = _borde_delgado()

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Hoja resumen
    ws_res = wb.create_sheet("Resumen Cruce SAP")
    total_sap    = len(df_sap)
    cruzados_sap = (df_sap["estado_cruce"] == "Cruzado").sum()
    sin_cruce_s  = (df_sap["estado_cruce"] == "Sin cruce").sum()
    debiles_sap  = (df_sap["estado_cruce"] == "Cruce débil").sum()

    ws_res.append(["Concepto", "Cantidad"])
    ws_res.append(["Total registros SAP", total_sap])
    ws_res.append(["Cruzados", cruzados_sap])
    ws_res.append(["Sin cruce", sin_cruce_s])
    ws_res.append(["Cruce débil", debiles_sap])
    pct = (cruzados_sap / max(total_sap, 1)) * 100
    ws_res.append(["% Cruzados", f"{pct:.1f}%"])

    wb.save(out_path)
    progreso("auxiliar_cruzado", 100, f"Auxiliar SAP marcado guardado: {out_path.name}")
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
# PASO 8 — ESCRITO WORD PARA EL SAT
# ══════════════════════════════════════════════════════════════════════════════

MACHOTE_TEXTO = """{{EMPRESA}}
{{DOMICILIO}}
R.F.C.: {{RFC_EMPRESA}}

{{FECHA_ESCRITO}}

Asunto: Solicitud de devolución del saldo a favor de Impuesto al Valor Agregado Convencional en cantidad de {{MONTO}} correspondiente al periodo de {{PERIODO}}.

R.F.C.: {{RFC_EMPRESA}}
Administración General de Grandes Contribuyentes
Administración Central de Fiscalización a Grandes Contribuyentes Diversos
Avenida Hidalgo No. 77, Col. Guerrero, Alcaldía Cuauhtémoc, 06300, Ciudad de México

P R E S E N T E

{{REP_LEGAL}} (R.F.C.: {{RFC_REP}}) apoderado legal de {{EMPRESA}} (R.F.C.: {{RFC_EMPRESA}}), personalidad que acredito mediante copia del instrumento notarial que se acompaña al presente como Anexo 1; igualmente se acompaña como parte del Anexo 1, copia de mi identificación oficial, señalando como domicilio para oír y recibir toda clase de notificaciones a que este asunto se refieran en {{DOMICILIO}}, y autorizando en los términos del artículo 19 del Código Fiscal de la Federación vigente a la fecha a los Sres.: {{AUTORIZADOS}}, atentamente comparezco y expongo que:

A N T E C E D E N T E S

I. {{EMPRESA}} (en adelante "mi representada"), es una sociedad debidamente constituida al amparo de las leyes de la República Mexicana que siempre ha cumplido con sus obligaciones fiscales.

II. El motivo o circunstancia que originó el saldo a favor de IVA correspondiente al periodo solicitado en devolución deriva en que mi representada tiene erogaciones cuyos proveedores trasladan IVA a la tasa del 16%, mientras que sus ingresos se encuentran gravados a tasa 0% o exentos, generando un saldo a favor acreditable conforme al Artículo 5 de la Ley del Impuesto al Valor Agregado.

Se proporciona como Anexo 2 carpeta electrónica que contiene copia del detalle de la declaración de IVA y su acuse de presentación, así como copia de la Declaración Informativa de Operaciones con Terceros "DIOT" correspondiente al periodo de {{PERIODO}}.

III. Se proporciona como Anexo 3 carátula del estado de cuenta de mi representada a efecto de que esa H. Administración pueda cotejar la cuenta CLABE de la cuenta bancaria a la que mi representada solicita que le sea efectuada la devolución del IVA, cuyo número CLABE es {{CLABE}}.

IV. A efectos de que esa H. Administración se encuentre en posibilidad de verificar la procedencia de la devolución del saldo a favor de IVA correspondiente al periodo de {{PERIODO}}, mi representada proporciona la información y documentación correspondiente, consistente en:

1. Como Anexo 4 papel de trabajo en formato electrónico Excel denominado "Integración IVA Acreditable" en el cual se integra el 100% de las erogaciones efectuadas por mi representada durante el periodo solicitado en devolución, relacionando entre otros datos: nombre del proveedor, folio fiscal (UUID) del Comprobante Fiscal Digital por Internet (CFDI), descripción del bien o servicio adquiridos, UUID del CFDI con complemento de pago y fecha de pago, debidamente referenciado con los estados de cuenta bancarios mediante la clave de cruce de auditoría asignada a cada operación, de los cuales se proporciona copia como Anexo 4.1. El reporte de integración fue generado procesando {{NUM_CFDIS}} CFDIs con Complemento de Pago correspondientes al periodo, de los cuales {{TOTAL_FACTURAS}} cuentan con cruce de auditoría confirmado tanto en el estado de cuenta bancario como en el auxiliar contable.

2. Como Anexo 4.2, respecto de los principales proveedores de mi representada en cuanto a monto se refiere, se proporciona carpeta en formato electrónico por cada proveedor, que contiene:
   a. Papel de trabajo de la integración de las principales operaciones en cuanto a monto se refiere, pagadas en el periodo solicitado en devolución, debidamente referenciadas con los estados de cuenta bancarios y su documentación comprobatoria.
   b. Copia de los CFDI correspondientes a las operaciones realizadas y pagadas durante el periodo solicitado en devolución.
   c. Copia de los CFDI con complemento de pagos, correspondientes a los pagos efectuados a dichos proveedores durante el periodo solicitado en devolución.

3. Como Anexo 4.3 el Auxiliar contable del IVA Acreditable efectivamente pagado en formato Excel emitido directamente del sistema contable SAP, donde es posible apreciar los cargos y abonos correspondientes al periodo solicitado en devolución, así como los identificadores de cruce de auditoría que vinculan cada registro contable con el CFDI con Complemento de Pago y el movimiento bancario correspondiente.

4. Mi representada manifiesta que los recursos económicos con los cuales se realizaron los gastos y erogaciones durante el periodo solicitado en devolución provienen de la operación normal de su actividad.

Índice de Anexos:

ANEXO 1     Poder e identificación oficial del representante legal
ANEXO 2     Declaración de IVA y DIOT del periodo {{PERIODO}}
ANEXO 3     Estado de cuenta bancario (carátula con CLABE)
ANEXO 4     Integración IVA Acreditable (Excel con cruces de auditoría)
ANEXO 4.1   Estados de cuenta bancarios cruzados — IVA acreditable
ANEXO 4.2   Integración principales proveedores
ANEXO 4.3   Auxiliar contable IVA acreditable (SAP) con marcas de cruce
ANEXO 5     Folio de solicitud SAT: {{FOLIO_SAT}}

Por lo anteriormente expuesto, solicito a esa H. Administración, lo siguiente:

PRIMERO. Se tenga por reconocida mi personalidad, acreditada mediante copia del instrumento notarial que se señala en el cuerpo del presente escrito.

SEGUNDO. Se valore la información y documentación que se proporciona mediante el presente escrito, relacionada al saldo a favor de Impuesto al Valor Agregado (IVA) correspondiente a {{PERIODO}}.

TERCERO. Se efectúe la devolución en cantidad de {{MONTO}} ({{MONTO_LETRAS}}) correspondiente al saldo a favor por concepto de Impuesto al Valor Agregado Convencional del mes {{PERIODO_MES}} {{ANIO}}.


{{EMPRESA}}

_________________________________
{{REP_LEGAL}}
Apoderado Legal
R.F.C.: {{RFC_REP}}
"""

MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}


def _monto_letras(monto: float) -> str:
    """Convierte monto a letras en español mayúsculas."""
    entero   = int(monto)
    centavos = round((monto - entero) * 100)
    texto    = num2words(entero, lang="es").upper()
    return f"{texto} PESOS {centavos:02d}/100 M.N."


def _reemplazar_en_parrafo(parrafo, marcadores: dict):
    """Reemplaza marcadores en un párrafo conservando el formato original."""
    for run in parrafo.runs:
        for marca, valor in marcadores.items():
            if marca in run.text:
                run.text = run.text.replace(marca, str(valor))
    # También manejar marcadores divididos entre runs
    texto_completo = "".join(run.text for run in parrafo.runs)
    for marca in marcadores:
        if marca in texto_completo:
            texto_completo = texto_completo.replace(marca, str(marcadores[marca]))
    if any(m in "".join(r.text for r in parrafo.runs) for m in marcadores):
        pass  # ya reemplazado en runs
    # Reemplazar en texto completo del párrafo (caso de runs partidos)
    if parrafo.runs:
        texto_nuevo = texto_completo
        if texto_nuevo != "".join(r.text for r in parrafo.runs):
            for run in parrafo.runs[1:]:
                run.text = ""
            parrafo.runs[0].text = texto_nuevo


def generar_escrito_word(cfg: dict, registros: list,
                          base_dir: Path, periodo_str: str) -> Path:
    """Genera el escrito de devolución de IVA en formato Word."""
    out_path  = base_dir / "output" / f"escrito_devolucion_IVA_{periodo_str}.docx"
    machote_p = base_dir / "input" / "machote" / "escrito_machote.docx"

    # Calcular datos del reporte
    iva_confirmado  = sum(r["iva16_mxn"] for r in registros
                         if r["cruce_edo_cuenta"] or r["cruce_sap"])
    total_cfdis     = len(set(r["uuid_cp"] for r in registros))
    cruce_completos = sum(1 for r in registros if r["cruce_edo_cuenta"] and r["cruce_sap"])

    # Determinar periodo
    # periodo_str formato YYYYMM
    try:
        año  = int(periodo_str[:4])
        mes  = int(periodo_str[4:6])
        mes_txt = MESES_ES.get(mes, str(mes))
    except Exception:
        año  = datetime.datetime.now().year
        mes_txt = "periodo"
        mes  = datetime.datetime.now().month

    hoy     = datetime.datetime.now()
    dia     = hoy.day
    mes_hoy = MESES_ES.get(hoy.month, "")
    año_hoy = hoy.year

    # Verificar campos faltantes
    campos_obligatorios = ["empresa", "rfc", "domicilio", "clabe",
                           "rep_legal", "rfc_rep", "autorizados"]
    for campo in campos_obligatorios:
        if not cfg.get(campo, "").strip():
            cfg[campo] = f"[COMPLETAR: {campo.upper()}]"

    folio_sat = cfg.get("folio_sat", "").strip() or "[FOLIO SOLICITUD SAT]"

    marcadores = {
        "{{EMPRESA}}":         cfg["empresa"],
        "{{RFC_EMPRESA}}":     cfg["rfc"],
        "{{DOMICILIO}}":       cfg["domicilio"],
        "{{CLABE}}":           cfg["clabe"],
        "{{REP_LEGAL}}":       cfg["rep_legal"],
        "{{RFC_REP}}":         cfg["rfc_rep"],
        "{{AUTORIZADOS}}":     cfg["autorizados"],
        "{{PERIODO}}":         f"{mes_txt} {año}",
        "{{PERIODO_MES}}":     mes_txt,
        "{{ANIO}}":            str(año),
        "{{MONTO}}":           f"${iva_confirmado:,.2f}",
        "{{MONTO_LETRAS}}":    _monto_letras(iva_confirmado),
        "{{FECHA_ESCRITO}}":   f"Ciudad de México a, {dia} de {mes_hoy} de {año_hoy}",
        "{{TOTAL_FACTURAS}}":  str(cruce_completos),
        "{{NUM_CFDIS}}":       str(total_cfdis),
        "{{FOLIO_SAT}}":       folio_sat,
    }

    # Generar siempre desde el texto integrado (no usar machote externo)
    doc = Document()
    sec = doc.sections[0]
    sec.page_width  = Cm(21.59)
    sec.page_height = Cm(27.94)
    sec.left_margin = sec.right_margin = sec.top_margin = sec.bottom_margin = Cm(2.5)

    texto_final = MACHOTE_TEXTO
    for marca, valor in marcadores.items():
        texto_final = texto_final.replace(marca, str(valor))

    for linea in texto_final.split("\n"):
        p = doc.add_paragraph(linea)
        p.alignment = WD_ALIGN_PARAGRAPH.JUSTIFY
        for run in p.runs:
            run.font.name = "Times New Roman"
            run.font.size = Pt(12)
        p.paragraph_format.space_after = Pt(0)
        p.paragraph_format.line_spacing = Pt(18)

    doc.save(str(out_path))
    progreso("word", 100, f"Escrito Word guardado: {out_path.name}")
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
# ANÁLISIS DE RIESGOS CON IA (Claude + Gemini en paralelo)
# ══════════════════════════════════════════════════════════════════════════════

_PROMPT_RIESGOS = """Eres un experto en auditoría fiscal mexicana especializado en IVA y devoluciones ante el SAT.

Analiza las siguientes operaciones de un contribuyente y evalúa el RIESGO DE RECHAZO del IVA acreditable por parte del SAT.

CONTRIBUYENTE: {empresa}
RFC: {rfc}
PERIODO: {periodo}

PROVEEDORES Y OPERACIONES (CFDIs tipo P — Complemento de Pago):
{operaciones}

Para cada proveedor, evalúa el riesgo considerando:
- Materialidad de las operaciones (Art. 5-A CFF, criterios de la PRODECON)
- Requisitos de deducibilidad (Art. 27 LISR, Art. 5 LIVA)
- Forma de pago (efectivo = mayor riesgo)
- RFC del proveedor (posible presencia en listas negras del SAT: EFOS/EDOS)
- Consistencia de los datos
- Jurisprudencias del TFJFA/SCJN en materia de IVA

Responde ÚNICAMENTE con un JSON válido con exactamente este formato (sin markdown, sin texto adicional):
{{
  "analisis": [
    {{
      "rfc_proveedor": "RFC aquí",
      "nombre_proveedor": "Nombre aquí",
      "nivel_riesgo": "BAJO|MEDIO|ALTO|CRÍTICO",
      "iva_en_riesgo": 12345.67,
      "factores_riesgo": ["factor 1", "factor 2"],
      "criterios_sat": ["Art. X LIVA", "Art. Y CFF"],
      "jurisprudencias": ["Tesis/jurisprudencia relevante"],
      "recomendaciones": ["Acción recomendada 1"],
      "documentacion_requerida": ["Documento 1", "Documento 2"]
    }}
  ],
  "resumen_general": "Resumen ejecutivo del riesgo global",
  "alertas_criticas": ["Alerta importante 1"]
}}"""


def _agrupar_proveedores(registros: list) -> list:
    """Agrupa los registros por RFC emisor para análisis de riesgos."""
    from collections import defaultdict
    grupos: dict = defaultdict(lambda: {
        "rfc": "", "nombre": "", "iva_total": 0.0, "monto_total": 0.0,
        "num_ops": 0, "con_cruce_banco": 0, "con_cruce_sap": 0,
        "formas_pago": set(), "tiene_efectivo": False,
    })
    for r in registros:
        rfc = r.get("rfc_emisor", "SIN_RFC")
        g = grupos[rfc]
        g["rfc"]   = rfc
        g["nombre"] = r.get("nombre_emisor", "")
        g["iva_total"]   += r.get("iva16_mxn", 0.0)
        g["monto_total"] += r.get("importe_pagado_mxn", 0.0)
        g["num_ops"] += 1
        if r.get("cruce_edo_cuenta"):
            g["con_cruce_banco"] += 1
        if r.get("cruce_sap"):
            g["con_cruce_sap"] += 1
        fp = r.get("forma_pago", "")
        g["formas_pago"].add(fp)
        if fp == "01":  # Efectivo
            g["tiene_efectivo"] = True
    # Serializar sets para JSON
    result = []
    for rfc, g in grupos.items():
        g["formas_pago"] = list(g["formas_pago"])
        sin_cruce = g["num_ops"] - max(g["con_cruce_banco"], g["con_cruce_sap"])
        g["sin_cruce"] = max(0, sin_cruce)
        result.append(g)
    return result


def _llamar_claude(prompt: str, api_key: str) -> dict:
    """Llama a la API de Claude y retorna el JSON de respuesta."""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        texto = msg.content[0].text.strip()
        # Limpiar posible markdown
        if texto.startswith("```"):
            texto = re.sub(r"^```[a-z]*\n?", "", texto)
            texto = re.sub(r"\n?```$", "", texto)
        return json.loads(texto)
    except Exception as exc:
        return {"error": str(exc), "analisis": [], "resumen_general": "",
                "alertas_criticas": []}


def _llamar_gemini(prompt: str, api_key: str) -> dict:
    """Llama a la API de Gemini y retorna el JSON de respuesta."""
    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel(
            "gemini-1.5-pro",
            generation_config={"response_mime_type": "application/json"},
        )
        resp = model.generate_content(prompt)
        return json.loads(resp.text)
    except Exception as exc:
        return {"error": str(exc), "analisis": [], "resumen_general": "",
                "alertas_criticas": []}


_NIVEL_ORDEN = {"BAJO": 0, "MEDIO": 1, "ALTO": 2, "CRÍTICO": 3}


def _combinar_analisis(lista_claude: list, lista_gemini: list) -> list:
    """
    Combina los análisis de Claude y Gemini.
    Para cada proveedor, toma el nivel de riesgo más alto y une factores.
    """
    gemini_by_rfc = {g["rfc_proveedor"]: g for g in lista_gemini
                     if "rfc_proveedor" in g}
    combinado = []
    for c in lista_claude:
        rfc = c.get("rfc_proveedor", "")
        g   = gemini_by_rfc.get(rfc, {})
        niv_c = c.get("nivel_riesgo", "BAJO")
        niv_g = g.get("nivel_riesgo", "BAJO")
        nivel = niv_c if _NIVEL_ORDEN.get(niv_c, 0) >= _NIVEL_ORDEN.get(niv_g, 0) else niv_g

        factores = list(dict.fromkeys(
            c.get("factores_riesgo", []) + g.get("factores_riesgo", [])
        ))
        criterios = list(dict.fromkeys(
            c.get("criterios_sat", []) + g.get("criterios_sat", [])
        ))
        jurisprudencias = list(dict.fromkeys(
            c.get("jurisprudencias", []) + g.get("jurisprudencias", [])
        ))
        recom = list(dict.fromkeys(
            c.get("recomendaciones", []) + g.get("recomendaciones", [])
        ))
        docs = list(dict.fromkeys(
            c.get("documentacion_requerida", []) + g.get("documentacion_requerida", [])
        ))
        combinado.append({
            "rfc_proveedor":         rfc,
            "nombre_proveedor":      c.get("nombre_proveedor", ""),
            "nivel_riesgo":          nivel,
            "nivel_claude":          niv_c,
            "nivel_gemini":          niv_g,
            "iva_en_riesgo":         c.get("iva_en_riesgo", 0),
            "factores_riesgo":       factores,
            "criterios_sat":         criterios,
            "jurisprudencias":       jurisprudencias,
            "recomendaciones":       recom,
            "documentacion_requerida": docs,
        })
    # Añadir los que solo están en Gemini
    rfcs_claude = {c["rfc_proveedor"] for c in lista_claude}
    for g in lista_gemini:
        if g.get("rfc_proveedor") not in rfcs_claude:
            g["nivel_claude"] = ""
            g["nivel_gemini"] = g.get("nivel_riesgo", "BAJO")
            combinado.append(g)
    return combinado


def analizar_riesgos_ia(registros_pago: list, cfg: dict) -> dict | None:
    """
    Analiza el riesgo de rechazo por IVA usando Claude y Gemini en paralelo.
    Retorna dict con 'analisis', 'resumen_general', 'alertas_criticas',
    o None si no hay API keys configuradas.
    """
    import threading as _thr

    anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "")
    gemini_key    = os.environ.get("GEMINI_API_KEY", "")

    if not anthropic_key and not gemini_key:
        progreso("riesgos", 0,
                 "ADVERTENCIA: ANTHROPIC_API_KEY y GEMINI_API_KEY no configuradas — omitiendo análisis IA")
        return None

    if not registros_pago:
        progreso("riesgos", 0, "Sin registros de pago para analizar riesgos")
        return None

    proveedores = _agrupar_proveedores(registros_pago)
    empresa  = cfg.get("empresa", "")
    rfc_emp  = cfg.get("rfc", "")
    periodo  = cfg.get("periodo_str", "")

    # Serializar proveedores para el prompt (limitar a top 50 por IVA)
    top_prov = sorted(proveedores, key=lambda x: x["iva_total"], reverse=True)[:50]
    ops_txt = json.dumps(top_prov, ensure_ascii=False, indent=2)
    prompt = _PROMPT_RIESGOS.format(
        empresa=empresa, rfc=rfc_emp, periodo=periodo, operaciones=ops_txt
    )

    resultado_claude: dict = {}
    resultado_gemini: dict = {}
    errores: list = []

    def _run_claude():
        if not anthropic_key:
            return
        progreso("riesgos", 20, "Consultando Claude...")
        r = _llamar_claude(prompt, anthropic_key)
        resultado_claude.update(r)
        if "error" in r:
            errores.append(f"Claude: {r['error']}")
        else:
            progreso("riesgos", 50, "Respuesta Claude recibida")

    def _run_gemini():
        if not gemini_key:
            return
        progreso("riesgos", 25, "Consultando Gemini...")
        r = _llamar_gemini(prompt, gemini_key)
        resultado_gemini.update(r)
        if "error" in r:
            errores.append(f"Gemini: {r['error']}")
        else:
            progreso("riesgos", 55, "Respuesta Gemini recibida")

    t1 = _thr.Thread(target=_run_claude)
    t2 = _thr.Thread(target=_run_gemini)
    t1.start(); t2.start()
    t1.join(); t2.join()

    for err in errores:
        print(f"ADVERTENCIA IA: {err}", flush=True)

    # Si solo hay una fuente, usarla directamente
    if resultado_claude and not resultado_gemini.get("analisis"):
        combinado = resultado_claude.get("analisis", [])
        for c in combinado:
            c["nivel_claude"] = c.get("nivel_riesgo", "BAJO")
            c["nivel_gemini"] = "—"
    elif resultado_gemini and not resultado_claude.get("analisis"):
        combinado = resultado_gemini.get("analisis", [])
        for c in combinado:
            c["nivel_gemini"] = c.get("nivel_riesgo", "BAJO")
            c["nivel_claude"] = "—"
    else:
        combinado = _combinar_analisis(
            resultado_claude.get("analisis", []),
            resultado_gemini.get("analisis", []),
        )

    # Ordenar por nivel de riesgo descendente
    combinado.sort(key=lambda x: _NIVEL_ORDEN.get(x.get("nivel_riesgo", "BAJO"), 0),
                   reverse=True)

    alertas = list(dict.fromkeys(
        resultado_claude.get("alertas_criticas", []) +
        resultado_gemini.get("alertas_criticas", [])
    ))
    resumen = (resultado_claude.get("resumen_general") or
               resultado_gemini.get("resumen_general") or "")

    return {"analisis": combinado, "resumen_general": resumen,
            "alertas_criticas": alertas}


_COLOR_BAJO     = "E2EFDA"
_COLOR_MEDIO    = "FFF2CC"
_COLOR_ALTO     = "FCE4D6"
_COLOR_CRITICO  = "FF0000"
_FG_CRITICO     = "FFFFFF"


def generar_reporte_riesgos(registros_pago: list, analisis: dict,
                             base_dir: Path, periodo_str: str) -> Path:
    """Genera el Excel de análisis de riesgos de rechazo por IVA."""
    out_path = base_dir / "output" / f"reporte_riesgos_{periodo_str}.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── HOJA 1: Análisis por Proveedor ────────────────────────────────────
    ws1 = wb.create_sheet("Análisis por Proveedor")
    enc1 = [
        "RFC Proveedor", "Nombre Proveedor", "IVA en Riesgo",
        "Nivel Riesgo", "Claude", "Gemini",
        "Factores de Riesgo", "Criterios SAT", "Jurisprudencias",
        "Recomendaciones", "Documentación Requerida",
    ]
    for col, e in enumerate(enc1, 1):
        c = ws1.cell(row=1, column=col, value=e)
        c.fill = _fill("1F4E79"); c.font = _font(bold=True, color="FFFFFF")
        c.border = _borde_delgado(); c.alignment = Alignment(horizontal="center", wrap_text=True)

    _COLOR_NIVEL = {
        "BAJO": (_COLOR_BAJO, "375623"),
        "MEDIO": (_COLOR_MEDIO, "7F6000"),
        "ALTO": (_COLOR_ALTO, "C00000"),
        "CRÍTICO": (_COLOR_CRITICO, _FG_CRITICO),
    }
    for i, a in enumerate(analisis.get("analisis", []), 2):
        nivel = a.get("nivel_riesgo", "BAJO")
        bg, fg = _COLOR_NIVEL.get(nivel, (_COLOR_BAJO, "375623"))
        datos = [
            a.get("rfc_proveedor", ""),
            a.get("nombre_proveedor", ""),
            a.get("iva_en_riesgo", 0),
            nivel,
            a.get("nivel_claude", ""),
            a.get("nivel_gemini", ""),
            " | ".join(a.get("factores_riesgo", [])),
            " | ".join(a.get("criterios_sat", [])),
            " | ".join(a.get("jurisprudencias", [])),
            " | ".join(a.get("recomendaciones", [])),
            " | ".join(a.get("documentacion_requerida", [])),
        ]
        for col, val in enumerate(datos, 1):
            c = ws1.cell(row=i, column=col, value=val)
            c.fill = _fill(bg); c.font = _font(color=fg); c.border = _borde_delgado()
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 3:
                c.number_format = '"$"#,##0.00'

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions
    anchos1 = [16, 32, 14, 10, 8, 8, 40, 30, 40, 40, 40]
    for col, w in enumerate(anchos1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── HOJA 2: Resumen Ejecutivo ─────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen Riesgos")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 22

    def _r_enc(txt):
        r = ws2.max_row + 1
        c = ws2.cell(row=r, column=1, value=txt)
        c.fill = _fill("1F4E79"); c.font = _font(bold=True, color="FFFFFF", size=11)
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

    def _r_fil(lbl, val):
        r = ws2.max_row + 1
        ws2.cell(row=r, column=1, value=lbl).font = _font(size=10)
        c = ws2.cell(row=r, column=2, value=val)
        c.font = _font(bold=True, size=10)
        c.alignment = Alignment(horizontal="right")

    lista_anal = analisis.get("analisis", [])
    conteos = {"BAJO": 0, "MEDIO": 0, "ALTO": 0, "CRÍTICO": 0}
    iva_por_nivel: dict = {"BAJO": 0.0, "MEDIO": 0.0, "ALTO": 0.0, "CRÍTICO": 0.0}
    for a in lista_anal:
        niv = a.get("nivel_riesgo", "BAJO")
        conteos[niv] = conteos.get(niv, 0) + 1
        iva_por_nivel[niv] = iva_por_nivel.get(niv, 0.0) + a.get("iva_en_riesgo", 0.0)

    _r_enc("ANÁLISIS DE RIESGOS — IVA DEVOLUCIONES SAT")
    _r_fil("Periodo", periodo_str)
    _r_fil("Generado", datetime.datetime.now().strftime("%d/%m/%Y %H:%M"))
    ws2.append([])
    _r_enc("DISTRIBUCIÓN POR NIVEL DE RIESGO")
    for niv in ["CRÍTICO", "ALTO", "MEDIO", "BAJO"]:
        _r_fil(f"  {niv} — proveedores", conteos.get(niv, 0))
        _r_fil(f"  {niv} — IVA en riesgo", f"${iva_por_nivel.get(niv, 0):,.2f}")
    ws2.append([])
    _r_enc("RESUMEN EJECUTIVO")
    resumen_txt = analisis.get("resumen_general", "")
    r_idx = ws2.max_row + 1
    c = ws2.cell(row=r_idx, column=1, value=resumen_txt)
    c.alignment = Alignment(wrap_text=True); c.font = _font(size=10)
    ws2.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=2)
    ws2.row_dimensions[r_idx].height = max(60, len(resumen_txt) // 3)
    ws2.append([])
    alertas = analisis.get("alertas_criticas", [])
    if alertas:
        _r_enc("ALERTAS CRÍTICAS")
        for alerta in alertas:
            r2 = ws2.max_row + 1
            c2 = ws2.cell(row=r2, column=1, value=f"⚠ {alerta}")
            c2.fill = _fill(_COLOR_ALTO); c2.font = _font(color="C00000")
            c2.alignment = Alignment(wrap_text=True)
            ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=2)

    # ── HOJA 3: Detalle por Operación ─────────────────────────────────────
    ws3 = wb.create_sheet("Detalle por Operación")
    enc3 = [
        "UUID Complemento", "Fecha Pago", "RFC Emisor", "Nombre Emisor",
        "IVA 16%", "Forma de Pago", "Cruce Banco", "Cruce SAP",
        "Nivel Riesgo (Proveedor)", "Principales Factores",
    ]
    for col, e in enumerate(enc3, 1):
        c = ws3.cell(row=1, column=col, value=e)
        c.fill = _fill("1F4E79"); c.font = _font(bold=True, color="FFFFFF")
        c.border = _borde_delgado()

    # Crear índice rfc→nivel para join rápido
    nivel_por_rfc = {a["rfc_proveedor"]: a for a in lista_anal}
    for i, r in enumerate(registros_pago, 2):
        rfc_e = r.get("rfc_emisor", "")
        inf = nivel_por_rfc.get(rfc_e, {})
        nivel = inf.get("nivel_riesgo", "SIN ANÁLISIS")
        bg, fg = _COLOR_NIVEL.get(nivel, ("FFFFFF", "000000"))
        factores = " | ".join(inf.get("factores_riesgo", [])[:2])
        datos3 = [
            r.get("uuid_cp", ""), r.get("fecha_pago", ""),
            rfc_e, r.get("nombre_emisor", ""),
            r.get("iva16_mxn", 0), r.get("forma_pago", ""),
            "✓" if r.get("cruce_edo_cuenta") else "✗",
            "✓" if r.get("cruce_sap") else "✗",
            nivel, factores,
        ]
        for col, val in enumerate(datos3, 1):
            c = ws3.cell(row=i, column=col, value=val)
            c.fill = _fill(bg); c.font = _font(color=fg); c.border = _borde_delgado()
            if col == 5: c.number_format = '"$"#,##0.00'

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions
    anchos3 = [36, 12, 14, 30, 12, 22, 8, 8, 16, 45]
    for col, w in enumerate(anchos3, 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    wb.save(str(out_path))
    progreso("riesgos", 100, f"Reporte de riesgos guardado: {out_path.name}")
    return out_path


# ══════════════════════════════════════════════════════════════════════════════
# MAIN — ORQUESTADOR
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # Directorio base: puede venir como argumento o ser el directorio actual
    if len(sys.argv) > 1:
        base_dir = Path(sys.argv[1])
    else:
        base_dir = Path(__file__).parent

    progreso("inicio", 0, f"Iniciando procesamiento en {base_dir}")

    # Cargar configuración
    cfg = cargar_config(base_dir)

    # Determinar periodo del mes actual (o del XML más reciente)
    hoy        = datetime.datetime.now()
    periodo_str = hoy.strftime("%Y%m")

    # ── PASO 2: Parseo CFDIs ────────────────────────────────────────────────
    # CFDIs de PAGO (proveedor es emisor) → IVA acreditable
    progreso("cfdi", 0, "Parseando CFDIs de pago...")
    registros_pago, errores_pago = parsear_cfdis(base_dir, carpeta="cfdi_pago",
                                                   tipo_cfdi="pago")
    # CFDIs de COBRO (nuestra empresa es emisora) → IVA trasladado
    progreso("cfdi", 50, "Parseando CFDIs de cobro...")
    registros_cobro, errores_cobro = parsear_cfdis(base_dir, carpeta="cfdi_cobro",
                                                    tipo_cfdi="cobro")
    errores_cfdi = errores_pago + errores_cobro
    # Para el procesamiento principal usamos los de pago
    registros = registros_pago
    todos_registros = registros_pago + registros_cobro
    if todos_registros:
        fechas = [r["fecha_pago"] for r in todos_registros if r["fecha_pago"]]
        if fechas:
            try:
                dt = datetime.datetime.strptime(sorted(fechas)[-1], "%Y-%m-%d")
                periodo_str = dt.strftime("%Y%m")
            except Exception:
                pass

    # ── PASO 3: Lectura estado de cuenta ───────────────────────────────────
    progreso("estado_cuenta", 5, "Leyendo PDFs del estado de cuenta...")
    movimientos = leer_estado_cuenta(base_dir)

    # ── PASO 4a: Auxiliar SAP IVA Acreditable (pagado) ─────────────────────
    progreso("auxiliar_sap", 5, "Leyendo auxiliar SAP IVA Acreditable...")
    df_sap_pago, col_map_pago, adv_pago = leer_auxiliar_sap(base_dir, "aux_pagado")
    for adv in adv_pago:
        print(f"ADVERTENCIA: {adv}", flush=True)

    # ── PASO 4b: Auxiliar SAP IVA Trasladado (cobrado) ──────────────────────
    progreso("auxiliar_sap", 20, "Leyendo auxiliar SAP IVA Trasladado...")
    df_sap_cobro, col_map_cobro, adv_cobro = leer_auxiliar_sap(base_dir, "aux_cobrado")
    for adv in adv_cobro:
        print(f"ADVERTENCIA: {adv}", flush=True)

    # ── Cruce PAGOS con banco ────────────────────────────────────────────────
    progreso("cruce_banco", 10, "Cruzando CFDIs de pago con banco...")
    cruce_banco_pago = cruzar_con_banco(registros_pago, movimientos, periodo_str)
    progreso("cruce_banco", 50,
             f"Cruce banco pagos: {cruce_banco_pago}/{len(registros_pago)}")

    # ── Cruce COBROS con banco ───────────────────────────────────────────────
    progreso("cruce_banco", 60, "Cruzando CFDIs de cobro con banco...")
    cruce_banco_cobro = cruzar_con_banco(registros_cobro, movimientos, periodo_str)
    progreso("cruce_banco", 100,
             f"Cruce banco cobros: {cruce_banco_cobro}/{len(registros_cobro)}")

    # ── Cruce PAGOS con SAP Acreditable ─────────────────────────────────────
    progreso("auxiliar_sap", 40, "Cruzando pagos con auxiliar SAP...")
    cruce_sap_pago = cruzar_con_sap(registros_pago, df_sap_pago,
                                     col_map_pago, periodo_str)

    # ── Cruce COBROS con SAP Trasladado ──────────────────────────────────────
    progreso("auxiliar_sap", 70, "Cruzando cobros con auxiliar SAP...")
    cruce_sap_cobro = cruzar_con_sap(registros_cobro, df_sap_cobro,
                                      col_map_cobro, periodo_str)

    # Variables legacy para compatibilidad con generadores
    registros = registros_pago
    df_sap    = df_sap_pago
    col_map   = col_map_pago

    # ── PASO 5: Excel reportes (uno por tipo) ───────────────────────────────
    progreso("excel", 10, "Generando reporte IVA Acreditable...")
    ruta_excel = generar_excel(registros_pago, movimientos, df_sap_pago,
                               errores_cfdi, base_dir, periodo_str,
                               tipo="acreditable")
    progreso("excel", 60, "Generando reporte IVA Trasladado...")
    ruta_excel_cobro = generar_excel(registros_cobro, movimientos, df_sap_cobro,
                                     errores_cfdi, base_dir, periodo_str,
                                     tipo="trasladado")

    # ── PASO 6: PDF marcado ─────────────────────────────────────────────────
    progreso("pdf", 10, "Marcando PDF del estado de cuenta...")
    ruta_pdf = marcar_pdf(movimientos, base_dir, periodo_str)

    # ── PASO 7: Auxiliares SAP cruzados ─────────────────────────────────────
    progreso("auxiliar_cruzado", 10, "Guardando auxiliar SAP Acreditable cruzado...")
    ruta_sap_out = guardar_auxiliar_sap_cruzado(df_sap_pago, base_dir,
                                                 periodo_str, tipo="acreditable")
    progreso("auxiliar_cruzado", 60, "Guardando auxiliar SAP Trasladado cruzado...")
    ruta_sap_cobro_out = guardar_auxiliar_sap_cruzado(df_sap_cobro, base_dir,
                                                       periodo_str, tipo="trasladado")

    # ── PASO 8: Escrito Word ─────────────────────────────────────────────────
    progreso("word", 10, "Generando escrito de devolución...")
    ruta_word = generar_escrito_word(cfg, registros, base_dir, periodo_str)

    # ── PASO 9: Análisis de Riesgos con IA ──────────────────────────────────
    ruta_riesgos = None
    cfg_riesgos = dict(cfg)
    cfg_riesgos["periodo_str"] = periodo_str
    progreso("riesgos", 5, "Iniciando análisis de riesgos con IA...")
    analisis_ia = analizar_riesgos_ia(registros_pago, cfg_riesgos)
    if analisis_ia is not None:
        progreso("riesgos", 80, "Generando reporte de riesgos...")
        ruta_riesgos = generar_reporte_riesgos(
            registros_pago, analisis_ia, base_dir, periodo_str
        )

    # ── RESUMEN FINAL ────────────────────────────────────────────────────────
    total_cfdis_cp  = len(set(r["uuid_cp"] for r in registros))
    total_facturas  = len(registros)
    iva_confirmado  = sum(r["iva16_mxn"] for r in registros
                         if r["cruce_edo_cuenta"] or r["cruce_sap"])
    cruces_banco_n  = sum(1 for r in registros if r["cruce_edo_cuenta"])
    cruces_sap_n    = sum(1 for r in registros if r["cruce_sap"])
    triple_n        = sum(1 for r in registros if r["cruce_edo_cuenta"] and r["cruce_sap"])
    sin_cruce_n     = sum(1 for r in registros if not r["cruce_edo_cuenta"] and not r["cruce_sap"])
    debiles_n       = sum(1 for r in registros if "débil" in r["metodo_cruce"].lower()
                          or "SOLO_MONTO" in r["metodo_cruce"])

    iva_acreditable = sum(r["iva16_mxn"] for r in registros_pago)
    iva_trasladado  = sum(r["iva16_mxn"] for r in registros_cobro)
    saldo_favor     = iva_trasladado - iva_acreditable

    resultado(total_cfdis_cp, iva_confirmado, triple_n, sin_cruce_n,
              iva_trasladado, iva_acreditable, saldo_favor)

    print("\n" + "-" * 60, flush=True)
    print(f"[OK] CFDIs Complemento de Pago procesados   : {total_cfdis_cp}", flush=True)
    print(f"[OK] Facturas relacionadas (DoctoRelacionado): {total_facturas}", flush=True)
    print(f"[OK] IVA 16% acreditable confirmado         : ${iva_confirmado:,.2f}", flush=True)
    print(f"[OK] Cruce con estado de cuenta             : {cruces_banco_n}/{total_facturas} "
          f"({cruces_banco_n/max(total_facturas,1)*100:.0f}%)", flush=True)
    print(f"[OK] Cruce con auxiliar SAP                 : {cruces_sap_n}/{total_facturas} "
          f"({cruces_sap_n/max(total_facturas,1)*100:.0f}%)", flush=True)
    print(f"[OK] Triple confirmacion (CFDI+banco+SAP)   : {triple_n} movimientos", flush=True)
    print(f"[!!] Sin cruce                              : {sin_cruce_n} movimientos", flush=True)
    print(f"[!!] Cruce debil (revisar manualmente)      : {debiles_n} movimientos", flush=True)
    print(f"\nArchivos generados en output/:", flush=True)
    print(f"   {ruta_excel.name if ruta_excel else 'N/A'}", flush=True)
    print(f"   {ruta_excel_cobro.name if ruta_excel_cobro else 'N/A'}", flush=True)
    print(f"   {ruta_sap_out.name if ruta_sap_out else 'N/A'}", flush=True)
    print(f"   {ruta_sap_cobro_out.name if ruta_sap_cobro_out else 'N/A'}", flush=True)
    print(f"   {ruta_pdf.name if ruta_pdf else 'N/A'}", flush=True)
    print(f"   {ruta_word.name}", flush=True)
    if ruta_riesgos:
        print(f"   {ruta_riesgos.name}", flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        error(f"{e}\n{traceback.format_exc()}")
        sys.exit(1)
